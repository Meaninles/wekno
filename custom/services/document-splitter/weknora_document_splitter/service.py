from __future__ import annotations

import copy
import csv
import hashlib
import html as html_lib
import io
import json
import math
import os
import posixpath
import shutil
import subprocess
import tempfile
import time
import zipfile
from dataclasses import dataclass
from email import policy as email_policy
from email.generator import BytesGenerator
from email.message import EmailMessage
from email.parser import BytesParser
from pathlib import Path
from typing import Any, Iterable, Iterator, Sequence
from urllib.parse import unquote, urlsplit

PLAN_VERSION = "documentsplit-v1"
_MB = 1024 * 1024
_SUPPORTED = {
    "pdf",
    "doc",
    "docx",
    "xls",
    "xlsx",
    "ppt",
    "pptx",
    "csv",
    "txt",
    "text",
    "md",
    "markdown",
    "json",
    "epub",
    "mhtml",
    "jpg",
    "jpeg",
    "png",
    "webp",
    "gif",
    "bmp",
    "tiff",
    "mp3",
    "m4a",
    "ogg",
    "flac",
    "wav",
}
_HARD_BYTES = {
    "docx": 10 * _MB,
    "xlsx": 10 * _MB,
    "pdf": 50 * _MB,
    "doc": 5 * _MB,
    "xls": 3 * _MB,
    "pptx": 50 * _MB,
    "ppt": 50 * _MB,
    "csv": 20 * _MB,
    "txt": 3 * _MB,
    "text": 3 * _MB,
    "md": 3 * _MB,
    "markdown": 3 * _MB,
    "json": 1 * _MB,
    "epub": 5 * _MB,
    "mhtml": 5 * _MB,
    "jpg": 5 * _MB,
    "jpeg": 5 * _MB,
    "png": 5 * _MB,
    "webp": 5 * _MB,
    "gif": 3 * _MB,
    "bmp": 2 * _MB,
    "tiff": 2 * _MB,
    "mp3": 100 * _MB,
    "m4a": 100 * _MB,
    "ogg": 100 * _MB,
    "flac": 100 * _MB,
    "wav": 5 * _MB,
}


class SplitFailure(RuntimeError):
    def __init__(self, code: str, message: str, retryable: bool = False):
        super().__init__(message)
        self.code = code
        self.retryable = retryable


@dataclass(frozen=True)
class SplitPolicy:
    max_source_bytes: int = 2 * 1024 * _MB
    max_parts: int = 10_000
    target_ratio: float = 0.75
    max_expansion_ratio: float = 12.0
    office_timeout_seconds: int = 1800
    audio_timeout_seconds: int = 1800
    image_tile_pixels: int = 16_000_000

    @classmethod
    def from_env(cls) -> "SplitPolicy":
        return cls(
            max_source_bytes=_positive_int_env(
                "CUSTOM_DOCUMENT_SPLIT_MAX_SOURCE_MB", 2048
            )
            * _MB,
            max_parts=_positive_int_env("CUSTOM_DOCUMENT_SPLIT_MAX_PARTS", 10_000),
            target_ratio=_ratio_env("CUSTOM_DOCUMENT_SPLIT_TARGET_RATIO", 0.75),
            max_expansion_ratio=_positive_float_env(
                "CUSTOM_DOCUMENT_SPLIT_MAX_EXPANSION_RATIO", 12.0
            ),
            office_timeout_seconds=_positive_int_env(
                "CUSTOM_DOCUMENT_SPLIT_OFFICE_TIMEOUT_SECONDS", 1800
            ),
            audio_timeout_seconds=_positive_int_env(
                "CUSTOM_DOCUMENT_SPLIT_AUDIO_TIMEOUT_SECONDS", 1800
            ),
            image_tile_pixels=_positive_int_env(
                "CUSTOM_DOCUMENT_SPLIT_IMAGE_TILE_PIXELS", 16_000_000
            ),
        )


@dataclass
class Part:
    path: Path
    file_type: str
    locator: dict[str, Any]
    metrics: dict[str, Any]


@dataclass(frozen=True)
class _MHTMLUnit:
    html: str
    source_index: int
    segment_index: int = 0
    segment_count: int = 1


def create_split_archive(
    *,
    source_path: Path,
    archive_path: Path,
    file_name: str,
    file_type: str,
    source_size: int,
    source_sha256: str,
    minimum_parts: int,
    target_ratio: float,
    policy: SplitPolicy,
) -> dict[str, Any]:
    ext = _normalize_ext(file_type or Path(file_name).suffix)
    if ext not in _SUPPORTED:
        raise SplitFailure("unsupported_format", f"unsupported split format: {ext}")
    if source_size <= 0:
        raise SplitFailure("empty_source", "empty documents cannot be split")
    if source_size > policy.max_source_bytes:
        raise SplitFailure(
            "source_too_large",
            f"source exceeds configured maximum of {policy.max_source_bytes} bytes",
        )
    ratio = target_ratio if 0.5 <= target_ratio <= 0.9 else policy.target_ratio
    part_dir = archive_path.parent / "parts"
    part_dir.mkdir(mode=0o700)

    parts, strategy = _split(
        source_path,
        part_dir,
        ext,
        max(2, minimum_parts),
        ratio,
        policy,
    )
    if not parts:
        raise SplitFailure("empty_split", "splitter produced no parseable parts")
    if len(parts) > policy.max_parts:
        raise SplitFailure(
            "too_many_parts",
            f"splitter produced {len(parts)} parts, limit is {policy.max_parts}",
        )

    total_part_bytes = 0
    manifest_parts: list[dict[str, Any]] = []
    for index, part in enumerate(parts):
        if not part.path.is_file() or part.path.is_symlink():
            raise SplitFailure("invalid_part", f"part {index} is not a regular file")
        size = part.path.stat().st_size
        if size <= 0:
            raise SplitFailure("empty_part", f"part {index} is empty")
        hard = _HARD_BYTES[part.file_type]
        if size > hard:
            raise SplitFailure(
                "part_limit_exceeded",
                f"part {index} is {size} bytes and exceeds {hard} bytes",
            )
        total_part_bytes += size
        name = f"part-{index + 1:06d}.{_output_extension(part.file_type)}"
        manifest_parts.append(
            {
                "index": index,
                "file_name": name,
                "file_type": part.file_type,
                "size_bytes": size,
                "sha256": _sha256_file(part.path),
                "locator": part.locator,
                "metrics": part.metrics,
            }
        )

    expansion_ratio = (
        policy.max_expansion_ratio
        if 1.0 <= policy.max_expansion_ratio <= 100.0
        else 12.0
    )
    max_expanded = max(
        512 * _MB,
        int(source_size * expansion_ratio),
    )
    if total_part_bytes > max_expanded:
        raise SplitFailure(
            "split_expansion_exceeded",
            f"split output expands to {total_part_bytes} bytes, limit is {max_expanded}",
        )

    manifest: dict[str, Any] = {
        "schema_version": 1,
        "planner_version": PLAN_VERSION,
        "created_unix_ms": int(time.time() * 1000),
        "strategy": strategy,
        "source": {
            "file_name": Path(file_name).name,
            "file_type": ext,
            "size_bytes": source_size,
            "sha256": source_sha256,
        },
        "target_ratio": ratio,
        "part_count": len(manifest_parts),
        "total_part_bytes": total_part_bytes,
        "parts": manifest_parts,
    }
    manifest_bytes = json.dumps(
        manifest, ensure_ascii=False, sort_keys=True, separators=(",", ":")
    ).encode("utf-8")

    with zipfile.ZipFile(
        archive_path, "x", compression=zipfile.ZIP_STORED, allowZip64=True
    ) as archive:
        archive.writestr("manifest.json", manifest_bytes)
        for metadata, part in zip(manifest_parts, parts, strict=True):
            archive.write(part.path, f"parts/{metadata['file_name']}")
    return manifest


def _split(
    source: Path,
    output: Path,
    ext: str,
    minimum_parts: int,
    ratio: float,
    policy: SplitPolicy,
) -> tuple[list[Part], str]:
    if ext == "pdf":
        return _split_pdf(source, output, minimum_parts, ratio), "pdf-pages"
    if ext in {"doc", "docx", "ppt", "pptx"}:
        pdf = _libreoffice_convert(
            source, output.parent / "normalized", "pdf", policy.office_timeout_seconds
        )
        parts = _split_pdf(pdf, output, minimum_parts, ratio)
        for part in parts:
            part.locator["normalized_from"] = ext
        return parts, f"{ext}-to-pdf-pages"
    if ext in {"xls", "xlsx"}:
        normalized = source
        if ext == "xls":
            normalized = _libreoffice_convert(
                source,
                output.parent / "normalized",
                "xlsx",
                policy.office_timeout_seconds,
            )
        return (
            _split_xlsx(normalized, output, minimum_parts, ratio, policy),
            f"{ext}-sheet-row-column-windows",
        )
    if ext == "csv":
        return _split_csv(source, output, minimum_parts, ratio), "csv-record-windows"
    if ext in {"txt", "text", "md", "markdown"}:
        return _split_text(source, output, ext, minimum_parts, ratio), "text-semantic-windows"
    if ext == "json":
        return _split_json(source, output, minimum_parts, ratio), "json-top-level-items"
    if ext == "epub":
        return _split_epub(source, output, minimum_parts, ratio), "epub-spine-items"
    if ext == "mhtml":
        return _split_mhtml(source, output, minimum_parts, ratio), "mhtml-dom-sections"
    if ext in {"jpg", "jpeg", "png", "webp", "gif", "bmp", "tiff"}:
        return _split_image(source, output, ext, minimum_parts, ratio, policy), "image-tiles"
    if ext in {"mp3", "m4a", "ogg", "flac", "wav"}:
        return _split_audio(source, output, ext, minimum_parts, ratio, policy), "audio-time-windows"
    raise SplitFailure("unsupported_format", f"unsupported split format: {ext}")


def _split_pdf(
    source: Path, output: Path, minimum_parts: int, ratio: float
) -> list[Part]:
    from pypdf import PdfReader, PdfWriter

    try:
        reader = PdfReader(str(source), strict=True)
    except Exception as exc:
        raise SplitFailure("invalid_pdf", f"PDF is malformed: {exc}") from exc
    if reader.is_encrypted:
        raise SplitFailure("encrypted_document", "encrypted PDF is not supported")
    page_count = len(reader.pages)
    if page_count == 0:
        raise SplitFailure("empty_document", "PDF contains no pages")

    hard = _HARD_BYTES["pdf"]
    target = max(_MB, int(hard * ratio))
    desired = max(minimum_parts, math.ceil(source.stat().st_size / target))
    desired = min(desired, page_count)
    page_weights: list[int] = []
    for page in reader.pages:
        weight = 1
        try:
            content = page.get_contents()
            if content is not None:
                data = content.get_data()
                weight = max(1, len(data))
        except Exception:
            pass
        page_weights.append(weight)
    ranges = _balanced_ranges(page_weights, desired)

    parts: list[Part] = []
    for index, (start, end) in enumerate(ranges):
        writer = PdfWriter()
        for page_index in range(start, end):
            writer.add_page(reader.pages[page_index])
        path = output / f"pdf-{index:06d}.pdf"
        with path.open("xb") as sink:
            writer.write(sink)
        parts.append(
            Part(
                path=path,
                file_type="pdf",
                locator={"kind": "pages", "page_start": start + 1, "page_end": end},
                metrics={"pages": end - start},
            )
        )
    return _resplit_oversized_pdf(parts, output, ratio)


def _resplit_oversized_pdf(parts: list[Part], output: Path, ratio: float) -> list[Part]:
    from pypdf import PdfReader, PdfWriter

    hard = _HARD_BYTES["pdf"]
    target = int(hard * ratio)
    result: list[Part] = []
    sequence = 0
    pending = list(parts)
    while pending:
        part = pending.pop(0)
        size = part.path.stat().st_size
        if size <= target:
            renamed = output / f"pdf-final-{sequence:06d}.pdf"
            sequence += 1
            if renamed != part.path:
                part.path.replace(renamed)
                part.path = renamed
            result.append(part)
            continue
        reader = PdfReader(str(part.path), strict=True)
        if len(reader.pages) <= 1:
            raise SplitFailure(
                "atomic_page_too_large",
                f"single PDF page remains {size} bytes and cannot be split losslessly",
            )
        midpoint = len(reader.pages) // 2
        page_start = int(part.locator["page_start"])
        children: list[Part] = []
        for child_index, (start, end) in enumerate(
            ((0, midpoint), (midpoint, len(reader.pages)))
        ):
            writer = PdfWriter()
            for page_index in range(start, end):
                writer.add_page(reader.pages[page_index])
            path = output / f"pdf-re-{time.time_ns()}-{child_index}.pdf"
            with path.open("xb") as sink:
                writer.write(sink)
            children.append(
                Part(
                    path,
                    "pdf",
                    {
                        "kind": "pages",
                        "page_start": page_start + start,
                        "page_end": page_start + end - 1,
                    },
                    {"pages": end - start},
                )
            )
        part.path.unlink(missing_ok=True)
        pending = children + pending
    return result


def _split_xlsx(
    source: Path,
    output: Path,
    minimum_parts: int,
    ratio: float,
    policy: SplitPolicy,
) -> list[Part]:
    from openpyxl import Workbook, load_workbook

    try:
        workbook = load_workbook(
            filename=source,
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except Exception as exc:
        raise SplitFailure("invalid_workbook", f"Excel workbook is malformed: {exc}") from exc

    target_bytes = int(_HARD_BYTES["xlsx"] * ratio)
    target_rows = max(1, int(30_000 * ratio))
    target_cells = max(1, int(300_000 * ratio))
    target_columns = max(1, int(100 * ratio))
    total_rows = sum(max(0, int(sheet.max_row or 0)) for sheet in workbook.worksheets)
    # Honour a size-derived minimum by reducing the source-row window up
    # front. Re-splitting generated workbooks would turn an injected
    # continuation header into a fake source row and corrupt coordinates.
    target_rows = min(
        target_rows,
        max(1, math.ceil(total_rows / max(1, minimum_parts))),
    )
    workbook.close()
    merge_ranges_by_sheet = _xlsx_merge_ranges_by_sheet(source)

    parts: list[Part] = []
    for sheet_index in range(_xlsx_sheet_count(source)):
        probe = load_workbook(source, read_only=True, data_only=True, keep_links=False)
        sheet = probe.worksheets[sheet_index]
        max_column = max(1, int(sheet.max_column or 1))
        sheet_name = sheet.title
        probe.close()
        sheet_merge_ranges = (
            merge_ranges_by_sheet[sheet_index]
            if sheet_index < len(merge_ranges_by_sheet)
            else []
        )
        anchor_column_count = min(3, max_column)
        for col_start in range(1, max_column + 1, target_columns):
            col_end = min(max_column, col_start + target_columns - 1)
            repeated_anchor_count = (
                anchor_column_count if col_start > anchor_column_count else 0
            )
            source_book = load_workbook(
                source, read_only=True, data_only=True, keep_links=False
            )
            source_sheet = source_book.worksheets[sheet_index]
            selected_merges = [
                merge_range
                for merge_range in sheet_merge_ranges
                if (
                    merge_range[2] <= col_end
                    and merge_range[3] >= col_start
                )
                or (
                    repeated_anchor_count
                    and merge_range[2] <= repeated_anchor_count
                    and merge_range[3] >= 1
                )
            ]
            merge_starts: dict[int, list[tuple[int, tuple[int, int, int, int]]]] = {}
            merge_ends: dict[int, list[int]] = {}
            for merge_index, merge_range in enumerate(selected_merges):
                min_row, max_row, _min_column, _max_column = merge_range
                merge_starts.setdefault(min_row, []).append(
                    (merge_index, merge_range)
                )
                merge_ends.setdefault(max_row, []).append(merge_index)
            active_merges: dict[int, tuple[int, int, int, int]] = {}
            merge_values: dict[int, Any] = {}
            header_values: tuple[Any, ...] | None = None
            header_row_index = 0
            writer: Workbook | None = None
            target_sheet = None
            path: Path | None = None
            part_row_start = 1
            rows_written = 0
            nonempty_cells = 0
            estimated_bytes = 0
            header_context = ""
            header_repeated = False
            last_source_row = 0

            def start_writer(row_start: int) -> None:
                nonlocal writer, target_sheet, path, part_row_start
                nonlocal header_context, header_repeated
                writer = Workbook(write_only=True)
                target_sheet = writer.create_sheet(title=_safe_sheet_title(sheet_name))
                path = output / f"xlsx-{len(parts):06d}-{time.time_ns()}.xlsx"
                part_row_start = row_start
                header_context = _format_column_header_context(
                    (header_values or ())[repeated_anchor_count:],
                    col_start,
                )
                if repeated_anchor_count:
                    anchor_context = _format_column_header_context(
                        (header_values or ())[:repeated_anchor_count], 1
                    )
                    if anchor_context:
                        header_context = (
                            f"{header_context}；关联关键列：{anchor_context}"
                            if header_context
                            else f"关联关键列：{anchor_context}"
                        )
                header_repeated = (
                    bool(header_context)
                    and header_row_index > 0
                    and row_start > header_row_index
                )
                if header_repeated:
                    target_sheet.append(list(header_values))

            def finish_writer(row_end: int) -> None:
                nonlocal writer, target_sheet, path, rows_written, nonempty_cells, estimated_bytes
                nonlocal last_source_row
                if writer is None or path is None or rows_written <= 0:
                    return
                writer.save(path)
                size = path.stat().st_size
                if size > _HARD_BYTES["xlsx"]:
                    path.unlink(missing_ok=True)
                    raise SplitFailure(
                        "atomic_excel_window_too_large",
                        f"Excel row window {sheet_name}!{part_row_start}:{row_end} exceeds 10MB",
                    )
                parts.append(
                    Part(
                        path=path,
                        file_type="xlsx",
                        locator={
                            "kind": "sheet_range",
                            "sheet": sheet_name,
                            "row_start": part_row_start,
                            "row_end": row_end,
                            "column_start": col_start,
                            "column_end": col_end,
                            "header_repeated": header_repeated,
                            "header_row": header_row_index,
                            "header_context": header_context,
                            "anchor_columns": (
                                list(range(1, repeated_anchor_count + 1))
                                if repeated_anchor_count
                                else []
                            ),
                            "merged_values_materialized": bool(selected_merges),
                        },
                        metrics={
                            "rows": rows_written,
                            "cells": nonempty_cells,
                            "estimated_uncompressed_bytes": estimated_bytes,
                            "file_bytes": size,
                        },
                    )
                )
                writer = None
                target_sheet = None
                path = None
                rows_written = 0
                nonempty_cells = 0
                estimated_bytes = 0
                last_source_row = 0

            for row_index, row in enumerate(
                source_sheet.iter_rows(
                    min_col=1 if repeated_anchor_count else col_start,
                    max_col=col_end,
                ),
                start=1,
            ):
                for merge_index, merge_range in merge_starts.get(
                    row_index, ()
                ):
                    active_merges[merge_index] = merge_range
                raw_values = [cell.value for cell in row]
                for merge_index, merge_range in active_merges.items():
                    min_row, _max_row, min_column, max_column_for_merge = (
                        merge_range
                    )
                    if row_index == min_row and min_column <= len(raw_values):
                        merge_values[merge_index] = raw_values[min_column - 1]
                    merged_value = merge_values.get(merge_index)
                    if merged_value is None:
                        continue
                    for column in range(
                        max(1, min_column),
                        min(col_end, max_column_for_merge) + 1,
                    ):
                        if raw_values[column - 1] is None:
                            raw_values[column - 1] = merged_value
                if repeated_anchor_count:
                    values = (
                        tuple(raw_values[:repeated_anchor_count])
                        + tuple(raw_values[col_start - 1 : col_end])
                    )
                else:
                    values = tuple(raw_values)
                for merge_index in merge_ends.get(row_index, ()):
                    active_merges.pop(merge_index, None)
                    merge_values.pop(merge_index, None)
                # Some production workbooks start with one or more visually
                # blank rows. Treat the first semantic row as the continuation
                # header so later physical windows retain the same column
                # meaning instead of losing it merely because A1 was empty.
                if header_values is None and any(
                    value is not None for value in values
                ):
                    header_values = values
                    header_row_index = row_index
                if not any(value is not None for value in values):
                    continue
                row_cells = sum(value is not None for value in values)
                row_bytes = sum(len(str(value).encode("utf-8")) for value in values if value is not None)
                if row_bytes > 8 * _MB:
                    source_book.close()
                    raise SplitFailure(
                        "atomic_excel_row_too_large",
                        f"Excel row {sheet_name}!{row_index} is too large to split safely",
                    )
                if writer is None:
                    start_writer(row_index)
                if (
                    rows_written > 0
                    and (
                        rows_written + 1 > target_rows
                        or nonempty_cells + row_cells > target_cells
                        or estimated_bytes + row_bytes > target_bytes
                    )
                ):
                    finish_writer(last_source_row)
                    start_writer(row_index)
                target_sheet.append(list(values))
                rows_written += 1
                nonempty_cells += row_cells
                estimated_bytes += row_bytes
                last_source_row = row_index
            finish_writer(last_source_row)
            source_book.close()

    if not parts:
        raise SplitFailure("empty_workbook", "Excel workbook contains no rows")
    sheet_order = {
        name: index
        for index, name in enumerate(_xlsx_sheet_names(source))
    }
    parts.sort(
        key=lambda part: (
            sheet_order.get(str(part.locator.get("sheet", "")), len(sheet_order)),
            int(part.locator.get("row_start", 0)),
            int(part.locator.get("column_start", 0)),
        )
    )
    return parts


def _xlsx_sheet_count(source: Path) -> int:
    from openpyxl import load_workbook

    workbook = load_workbook(source, read_only=True, data_only=True, keep_links=False)
    try:
        return len(workbook.worksheets)
    finally:
        workbook.close()


def _xlsx_sheet_names(source: Path) -> list[str]:
    from openpyxl import load_workbook

    workbook = load_workbook(source, read_only=True, data_only=True, keep_links=False)
    try:
        return [sheet.title for sheet in workbook.worksheets]
    finally:
        workbook.close()


def _xlsx_merge_ranges_by_sheet(
    source: Path,
) -> list[list[tuple[int, int, int, int]]]:
    """Read merge coordinates directly from OOXML without loading cell grids."""
    from xml.etree import ElementTree

    from openpyxl.utils.cell import range_boundaries

    try:
        with zipfile.ZipFile(source) as archive:
            workbook = ElementTree.fromstring(archive.read("xl/workbook.xml"))
            relationships = ElementTree.fromstring(
                archive.read("xl/_rels/workbook.xml.rels")
            )
            targets = {
                relation.attrib.get("Id", ""): relation.attrib.get("Target", "")
                for relation in relationships.iter()
                if relation.tag.rsplit("}", 1)[-1] == "Relationship"
            }
            sheet_paths: list[str] = []
            relationship_namespace = (
                "http://schemas.openxmlformats.org/officeDocument/"
                "2006/relationships"
            )
            for sheet in workbook.iter():
                if sheet.tag.rsplit("}", 1)[-1] != "sheet":
                    continue
                relationship_id = sheet.attrib.get(
                    f"{{{relationship_namespace}}}id", ""
                )
                target = targets.get(relationship_id, "")
                if target.startswith("/"):
                    normalized = target.lstrip("/")
                else:
                    normalized = posixpath.normpath(
                        posixpath.join("xl", target)
                    )
                if not normalized.startswith("xl/worksheets/"):
                    raise SplitFailure(
                        "invalid_workbook",
                        "Excel worksheet relationship points outside worksheets",
                    )
                sheet_paths.append(normalized)

            result: list[list[tuple[int, int, int, int]]] = []
            for sheet_path in sheet_paths:
                ranges: list[tuple[int, int, int, int]] = []
                with archive.open(sheet_path) as stream:
                    for _event, element in ElementTree.iterparse(
                        stream, events=("end",)
                    ):
                        if element.tag.rsplit("}", 1)[-1] == "mergeCell":
                            reference = element.attrib.get("ref", "")
                            try:
                                min_column, min_row, max_column, max_row = (
                                    range_boundaries(reference)
                                )
                            except Exception as exc:
                                raise SplitFailure(
                                    "invalid_workbook",
                                    f"invalid Excel merge range {reference!r}",
                                ) from exc
                            ranges.append(
                                (min_row, max_row, min_column, max_column)
                            )
                            if len(ranges) > 100_000:
                                raise SplitFailure(
                                    "too_many_merge_ranges",
                                    "Excel sheet contains more than 100,000 merge ranges",
                                )
                        element.clear()
                result.append(ranges)
            return result
    except SplitFailure:
        raise
    except Exception as exc:
        raise SplitFailure(
            "invalid_workbook",
            f"cannot inspect Excel merge ranges: {exc}",
        ) from exc


def _split_csv(
    source: Path, output: Path, minimum_parts: int, ratio: float
) -> list[Part]:
    encoding = _detect_text_encoding(source)
    dialect = _csv_dialect(source, encoding)
    target_bytes = int(_HARD_BYTES["csv"] * ratio)
    target_rows = max(1, int(100_000 * ratio))
    target_cells = max(1, int(300_000 * ratio))
    target_columns = max(1, int(100 * ratio))

    max_columns = 1
    row_count = 0
    with source.open("r", encoding=encoding, newline="") as stream:
        reader = csv.reader(stream, dialect)
        for row in reader:
            row_count += 1
            max_columns = max(max_columns, len(row))
            for field in row:
                if len(field.encode("utf-8")) > 512 * 1024:
                    raise SplitFailure(
                        "atomic_csv_field_too_large",
                        f"CSV row {row_count} contains a field larger than 512KB",
                    )
    row_target_from_minimum = max(1, math.ceil(row_count / minimum_parts))
    target_rows = min(target_rows, row_target_from_minimum)

    parts: list[Part] = []
    anchor_column_count = min(3, max_columns)
    for col_start in range(0, max_columns, target_columns):
        col_end = min(max_columns, col_start + target_columns)
        repeated_anchor_count = (
            anchor_column_count if col_start >= anchor_column_count else 0
        )
        with source.open("r", encoding=encoding, newline="") as stream:
            reader = csv.reader(stream, dialect)
            header: list[str] | None = None
            sink = None
            writer = None
            path = None
            start_row = 1
            written = 0
            cells = 0
            estimated = 0
            header_context = ""
            header_repeated = False

            def begin(row_index: int) -> None:
                nonlocal sink, writer, path, start_row
                nonlocal header_context, header_repeated
                path = output / f"csv-{len(parts):06d}-{time.time_ns()}.csv"
                sink = path.open("x", encoding="utf-8", newline="")
                writer = csv.writer(sink, dialect)
                start_row = row_index
                header_context = _format_column_header_context(
                    (header or [])[col_start:col_end], col_start + 1
                )
                if repeated_anchor_count:
                    anchor_context = _format_column_header_context(
                        (header or [])[:repeated_anchor_count], 1
                    )
                    if anchor_context:
                        header_context = (
                            f"{header_context}；关联关键列：{anchor_context}"
                            if header_context
                            else f"关联关键列：{anchor_context}"
                        )
                header_repeated = bool(header_context) and row_index > 1
                if header_repeated:
                    writer.writerow(
                        header[:repeated_anchor_count]
                        + header[col_start:col_end]
                    )

            def finish(row_end: int) -> None:
                nonlocal sink, writer, path, written, cells, estimated
                if sink is None or path is None or written <= 0:
                    return
                sink.flush()
                os.fsync(sink.fileno())
                sink.close()
                parts.append(
                    Part(
                        path,
                        "csv",
                        {
                            "kind": "record_range",
                            "row_start": start_row,
                            "row_end": row_end,
                            "column_start": col_start + 1,
                            "column_end": col_end,
                            "header_repeated": header_repeated,
                            "header_context": header_context,
                            "anchor_columns": (
                                list(range(1, repeated_anchor_count + 1))
                                if repeated_anchor_count
                                else []
                            ),
                        },
                        {"rows": written, "cells": cells, "file_bytes": path.stat().st_size},
                    )
                )
                sink = None
                writer = None
                path = None
                written = 0
                cells = 0
                estimated = 0

            for row_index, row in enumerate(reader, start=1):
                if row_index == 1:
                    header = row
                window = (
                    row[:repeated_anchor_count] + row[col_start:col_end]
                    if repeated_anchor_count
                    else row[col_start:col_end]
                )
                row_bytes = sum(len(field.encode("utf-8")) for field in window)
                if row_bytes > _MB:
                    raise SplitFailure(
                        "atomic_csv_row_too_large",
                        f"CSV row {row_index} exceeds the 1MB atomic row limit",
                    )
                if sink is None:
                    begin(row_index)
                if (
                    written > 0
                    and (
                        written + 1 > target_rows
                        or cells + len(window) > target_cells
                        or estimated + row_bytes > target_bytes
                    )
                ):
                    finish(row_index - 1)
                    begin(row_index)
                writer.writerow(window)
                written += 1
                cells += len(window)
                estimated += row_bytes
            finish(row_count)
    parts.sort(
        key=lambda part: (
            int(part.locator.get("row_start", 0)),
            int(part.locator.get("column_start", 0)),
        )
    )
    return parts


def _split_text(
    source: Path,
    output: Path,
    ext: str,
    minimum_parts: int,
    ratio: float,
) -> list[Part]:
    normalized_ext = "md" if ext in {"md", "markdown"} else "txt"
    target = int(_HARD_BYTES[ext] * ratio)
    size_target = min(target, max(64 * 1024, math.ceil(source.stat().st_size / minimum_parts)))
    encoding = _detect_text_encoding(source)
    parts: list[Part] = []
    active_headings: list[str] = []
    sink = None
    path = None
    start_line = 1
    start_character = 0
    end_line = 0
    end_character = 0
    written = 0
    source_units_written = 0

    def begin(line_number: int, character_start: int) -> None:
        nonlocal sink, path, start_line, start_character
        nonlocal written, source_units_written
        path = output / f"text-{len(parts):06d}.{normalized_ext}"
        sink = path.open("x", encoding="utf-8", newline="")
        start_line = line_number
        start_character = character_start
        written = 0
        source_units_written = 0
        if normalized_ext == "md" and active_headings:
            prefix = "\n".join(active_headings) + "\n\n"
            sink.write(prefix)
            written += len(prefix.encode("utf-8"))

    def finish() -> None:
        nonlocal sink, path, written, source_units_written
        if sink is None or path is None or source_units_written <= 0:
            return
        sink.flush()
        os.fsync(sink.fileno())
        sink.close()
        locator: dict[str, Any] = {
            "kind": "line_range",
            "line_start": start_line,
            "line_end": end_line,
        }
        if start_character > 0 or (
            start_line == end_line and end_character > 0
        ):
            locator["character_start"] = start_character
            locator["character_end"] = end_character
        parts.append(
            Part(
                path,
                normalized_ext,
                locator,
                {
                    "lines": end_line - start_line + 1,
                    "file_bytes": path.stat().st_size,
                },
            )
        )
        sink = None
        path = None
        written = 0
        source_units_written = 0

    with source.open("r", encoding=encoding, errors="strict", newline="") as stream:
        for line_number, line in enumerate(stream, start=1):
            character_offset = 0
            fragments = _semantic_text_fragments(line, size_target)
            for fragment in fragments:
                encoded = fragment.encode("utf-8")
                if len(encoded) > _HARD_BYTES[ext]:
                    raise SplitFailure(
                        "atomic_text_fragment_too_large",
                        f"text fragment on line {line_number} exceeds the parser hard limit",
                    )
                if sink is None:
                    begin(line_number, character_offset)
                if (
                    source_units_written > 0
                    and written + len(encoded) > size_target
                ):
                    finish()
                    begin(line_number, character_offset)
                sink.write(fragment)
                written += len(encoded)
                source_units_written += 1
                end_line = line_number
                character_offset += len(fragment)
                end_character = character_offset
            if normalized_ext == "md":
                stripped = line.lstrip()
                if stripped.startswith("#"):
                    level = len(stripped) - len(stripped.lstrip("#"))
                    if 1 <= level <= 6 and stripped[level : level + 1] == " ":
                        active_headings[:] = [
                            value
                            for value in active_headings
                            if len(value) - len(value.lstrip("#")) < level
                        ]
                        active_headings.append(stripped.rstrip())
    finish()
    return parts


def _semantic_text_fragments(value: str, byte_budget: int) -> list[str]:
    if len(value.encode("utf-8")) <= byte_budget:
        return [value]
    fragments: list[str] = []
    remaining = value
    boundaries = set("\n\r\t 。！？!?；;，,、")
    while remaining:
        used = 0
        end = 0
        preferred = 0
        for index, character in enumerate(remaining):
            size = len(character.encode("utf-8"))
            if used + size > byte_budget:
                break
            used += size
            end = index + 1
            if character in boundaries and used >= byte_budget // 2:
                preferred = end
        if end == 0:
            raise SplitFailure(
                "atomic_text_character_too_large",
                "one encoded text character exceeds the split byte budget",
            )
        cut = preferred or end
        fragments.append(remaining[:cut])
        remaining = remaining[cut:]
    return fragments


def _split_json(
    source: Path, output: Path, minimum_parts: int, ratio: float
) -> list[Part]:
    try:
        import ijson
    except ImportError as exc:
        raise SplitFailure(
            "splitter_dependency_missing", "streaming JSON splitter is unavailable"
        ) from exc

    target = min(
        int(_HARD_BYTES["json"] * ratio),
        max(64 * 1024, math.ceil(source.stat().st_size / minimum_parts)),
    )
    with source.open("rb") as probe:
        prefix = probe.read(64 * 1024).lstrip(b"\xef\xbb\xbf \t\r\n")
    if not prefix:
        raise SplitFailure("empty_document", "JSON source is empty")
    if _contains_multiple_json_values(source, ijson):
        return _split_ndjson(source, output, minimum_parts, ratio)
    top_array = prefix.startswith(b"[")
    top_object = prefix.startswith(b"{")
    if not top_array and not top_object:
        return _split_ndjson(source, output, minimum_parts, ratio)
    if (
        top_object
        and source.stat().st_size > _HARD_BYTES["json"]
        and _json_root_has_nested_container(source, ijson)
    ):
        return _split_json_path_records(
            source, output, minimum_parts, ratio, ijson
        )

    parts: list[Part] = []
    current: list[Any] | dict[str, Any]
    current = [] if top_array else {}
    current_bytes = 2
    item_start = 0
    item_end = -1

    def flush() -> None:
        nonlocal current, current_bytes, item_start
        if not current:
            return
        index = len(parts)
        path = output / f"json-{index:06d}.json"
        with path.open("x", encoding="utf-8", newline="\n") as sink:
            json.dump(current, sink, ensure_ascii=False, separators=(",", ":"))
        parts.append(
            Part(
                path,
                "json",
                {
                    "kind": "json_items",
                    "container_kind": "array" if top_array else "object",
                    "item_start": item_start,
                    "item_end": item_end,
                },
                {"items": len(current), "file_bytes": path.stat().st_size},
            )
        )
        item_start = item_end + 1
        current = [] if top_array else {}
        current_bytes = 2

    with source.open("rb") as stream:
        iterator: Iterable[Any]
        if top_array:
            iterator = ijson.items(stream, "item", use_float=True)
            for item_end, value in enumerate(iterator):
                encoded = json.dumps(
                    value, ensure_ascii=False, separators=(",", ":")
                ).encode("utf-8")
                if len(encoded) > _HARD_BYTES["json"] - 2:
                    raise SplitFailure(
                        "atomic_json_item_too_large",
                        f"JSON array item {item_end} exceeds the 1MB atomic limit",
                    )
                if current and current_bytes + len(encoded) + 1 > target:
                    flush()
                current.append(value)
                current_bytes += len(encoded) + 1
        else:
            iterator = ijson.kvitems(stream, "", use_float=True)
            for item_end, (key, value) in enumerate(iterator):
                encoded = json.dumps(
                    {key: value}, ensure_ascii=False, separators=(",", ":")
                ).encode("utf-8")
                if len(encoded) > _HARD_BYTES["json"] - 2:
                    raise SplitFailure(
                        "atomic_json_member_too_large",
                        f"JSON object member {key!r} exceeds the 1MB atomic limit",
                    )
                if current and current_bytes + len(encoded) + 1 > target:
                    flush()
                current[key] = value
                current_bytes += len(encoded) + 1
    flush()
    return parts


def _json_root_has_nested_container(source: Path, ijson_module: Any) -> bool:
    depth = 0
    try:
        with source.open("rb") as stream:
            for _prefix, event, _value in ijson_module.parse(
                stream, use_float=True
            ):
                if event in {"start_map", "start_array"}:
                    depth += 1
                    if depth == 2:
                        return True
                elif event in {"end_map", "end_array"}:
                    depth -= 1
    except Exception as exc:
        raise SplitFailure("invalid_json", f"JSON source is malformed: {exc}") from exc
    return False


def _split_json_path_records(
    source: Path,
    output: Path,
    minimum_parts: int,
    ratio: float,
    ijson_module: Any,
) -> list[Part]:
    target = min(
        int(_HARD_BYTES["json"] * ratio),
        max(64 * 1024, math.ceil(source.stat().st_size / minimum_parts)),
    )
    parts: list[Part] = []
    current: list[dict[str, Any]] = []
    current_bytes = len(b'{"__weknora_path_records":[]}')
    record_start = 0
    record_end = -1

    def flush() -> None:
        nonlocal current, current_bytes, record_start
        if not current:
            return
        path = output / f"json-path-{len(parts):06d}.json"
        payload = {"__weknora_path_records": current}
        with path.open("x", encoding="utf-8", newline="\n") as sink:
            json.dump(payload, sink, ensure_ascii=False, separators=(",", ":"))
        parts.append(
            Part(
                path,
                "json",
                {
                    "kind": "json_path_records",
                    "record_start": record_start,
                    "record_end": record_end,
                    "path_start": current[0]["path"],
                    "path_end": current[-1]["path"],
                },
                {"records": len(current), "file_bytes": path.stat().st_size},
            )
        )
        record_start = record_end + 1
        current = []
        current_bytes = len(b'{"__weknora_path_records":[]}')

    def append_record(record: dict[str, Any]) -> None:
        nonlocal current_bytes, record_end
        encoded = json.dumps(
            record, ensure_ascii=False, separators=(",", ":")
        ).encode("utf-8")
        if len(encoded) + 64 > _HARD_BYTES["json"]:
            raise SplitFailure(
                "atomic_json_scalar_too_large",
                f"JSON scalar at {record.get('path', '$')} exceeds the 1MB limit",
            )
        if current and current_bytes + len(encoded) + 1 > target:
            flush()
        current.append(record)
        current_bytes += len(encoded) + 1
        record_end += 1

    # Frames track exact array indexes and object keys, which ijson's prefix
    # alone cannot express (it reports every array element as ".item").
    frames: list[dict[str, Any]] = []

    def consume_path() -> str:
        if not frames:
            return "$"
        parent = frames[-1]
        parent["children"] += 1
        if parent["kind"] == "array":
            index = parent["next_index"]
            parent["next_index"] += 1
            return f"{parent['path']}[{index}]"
        key = parent.get("pending_key")
        parent["pending_key"] = None
        if key is None:
            raise SplitFailure("invalid_json", "JSON object value has no key")
        escaped = json.dumps(str(key), ensure_ascii=False)
        return f"{parent['path']}[{escaped}]"

    try:
        with source.open("rb") as stream:
            for _prefix, event, value in ijson_module.parse(
                stream, use_float=True
            ):
                if event == "map_key":
                    if not frames or frames[-1]["kind"] != "map":
                        raise SplitFailure("invalid_json", "JSON key is outside an object")
                    frames[-1]["pending_key"] = value
                    continue
                if event in {"start_map", "start_array"}:
                    value_path = consume_path()
                    frames.append(
                        {
                            "kind": "map" if event == "start_map" else "array",
                            "path": value_path,
                            "pending_key": None,
                            "next_index": 0,
                            "children": 0,
                        }
                    )
                    continue
                if event in {"end_map", "end_array"}:
                    if not frames:
                        raise SplitFailure("invalid_json", "JSON container stack underflow")
                    frame = frames.pop()
                    if frame["children"] == 0:
                        append_record(
                            {
                                "path": frame["path"],
                                "value": {} if frame["kind"] == "map" else [],
                            }
                        )
                    continue
                if event in {"string", "number", "boolean", "null"}:
                    value_path = consume_path()
                    if isinstance(value, str) and len(value.encode("utf-8")) > target // 2:
                        character_start = 0
                        for fragment in _semantic_text_fragments(
                            value, max(1024, target // 2)
                        ):
                            character_end = character_start + len(fragment)
                            append_record(
                                {
                                    "path": value_path,
                                    "value_fragment": fragment,
                                    "character_start": character_start,
                                    "character_end": character_end,
                                }
                            )
                            character_start = character_end
                    else:
                        append_record({"path": value_path, "value": value})
    except SplitFailure:
        raise
    except Exception as exc:
        raise SplitFailure("invalid_json", f"JSON source is malformed: {exc}") from exc
    flush()
    if not parts:
        raise SplitFailure("empty_document", "JSON source contains no values")
    return parts


def _contains_multiple_json_values(source: Path, ijson_module: Any) -> bool:
    roots = 0
    depth = 0
    try:
        with source.open("rb") as stream:
            for _prefix, event, _value in ijson_module.parse(
                stream, multiple_values=True, use_float=True
            ):
                if event in {"start_map", "start_array"}:
                    depth += 1
                    continue
                if event in {"end_map", "end_array"}:
                    depth -= 1
                    if depth == 0:
                        roots += 1
                elif depth == 0 and event in {
                    "string",
                    "number",
                    "boolean",
                    "null",
                }:
                    roots += 1
                if roots > 1:
                    return True
    except Exception as exc:
        raise SplitFailure("invalid_json", f"JSON source is malformed: {exc}") from exc
    return False


def _split_ndjson(
    source: Path, output: Path, minimum_parts: int, ratio: float
) -> list[Part]:
    encoding = _detect_text_encoding(source)
    target = min(
        int(_HARD_BYTES["json"] * ratio),
        max(64 * 1024, math.ceil(source.stat().st_size / minimum_parts)),
    )
    parts: list[Part] = []
    current: list[Any] = []
    current_bytes = 2
    item_start = 0
    item_end = -1

    def flush() -> None:
        nonlocal current, current_bytes, item_start
        if not current:
            return
        path = output / f"ndjson-{len(parts):06d}.json"
        with path.open("x", encoding="utf-8", newline="\n") as sink:
            json.dump(current, sink, ensure_ascii=False, separators=(",", ":"))
        parts.append(
            Part(
                path,
                "json",
                {
                    "kind": "json_items",
                    "content_kind": "ndjson",
                    "item_start": item_start,
                    "item_end": item_end,
                },
                {"items": len(current), "file_bytes": path.stat().st_size},
            )
        )
        item_start = item_end + 1
        current = []
        current_bytes = 2

    with source.open("r", encoding=encoding, errors="strict", newline="") as stream:
        for line_number, line in enumerate(stream, start=1):
            stripped = line.strip()
            if not stripped:
                continue
            try:
                value = json.loads(stripped)
            except json.JSONDecodeError as exc:
                raise SplitFailure(
                    "invalid_ndjson",
                    f"NDJSON line {line_number} is invalid: {exc}",
                ) from exc
            encoded = json.dumps(
                value, ensure_ascii=False, separators=(",", ":")
            ).encode("utf-8")
            if len(encoded) > _HARD_BYTES["json"] - 2:
                raise SplitFailure(
                    "atomic_json_item_too_large",
                    f"NDJSON item at line {line_number} exceeds the 1MB atomic limit",
                )
            if current and current_bytes + len(encoded) + 1 > target:
                flush()
            item_end += 1
            current.append(value)
            current_bytes += len(encoded) + 1
    flush()
    if not parts:
        raise SplitFailure("empty_document", "NDJSON source contains no records")
    return parts


def _split_epub(
    source: Path, output: Path, minimum_parts: int, ratio: float
) -> list[Part]:
    import ebooklib
    from ebooklib import epub

    try:
        book = epub.read_epub(str(source), options={"ignore_ncx": False})
    except Exception as exc:
        raise SplitFailure("invalid_epub", f"EPUB is malformed: {exc}") from exc
    # get_items_of_type() follows manifest order, which is not guaranteed to
    # be reading order. Prefer the EPUB spine and append any orphan readable
    # documents deterministically.
    document_by_id = {
        item.get_id(): item
        for item in book.get_items_of_type(ebooklib.ITEM_DOCUMENT)
    }
    documents = []
    seen_document_ids: set[str] = set()
    for spine_entry in book.spine:
        item_id = spine_entry[0] if isinstance(spine_entry, tuple) else spine_entry
        document = document_by_id.get(item_id)
        if document is None or item_id in seen_document_ids:
            continue
        if hasattr(document, "is_chapter") and not document.is_chapter():
            continue
        documents.append(document)
        seen_document_ids.add(item_id)
    for item_id, document in document_by_id.items():
        if item_id in seen_document_ids:
            continue
        if hasattr(document, "is_chapter") and not document.is_chapter():
            continue
        documents.append(document)
        seen_document_ids.add(item_id)
    if not documents:
        raise SplitFailure("empty_document", "EPUB contains no readable spine items")

    image_resources: dict[str, Any] = {}
    for item in book.get_items_of_type(ebooklib.ITEM_IMAGE):
        file_name = str(getattr(item, "file_name", "") or "").strip()
        if not file_name:
            continue
        image_resources[posixpath.normpath(file_name)] = item
    referenced_resources = [
        _epub_referenced_resources(document, image_resources)
        for document in documents
    ]
    target = int(_HARD_BYTES["epub"] * ratio)
    resource_sizes = {
        name: len(resource.get_content())
        for name, resource in image_resources.items()
    }

    # EPUB resources are commonly reused by many consecutive chapters. Adding
    # each chapter's referenced-image bytes to a scalar weight either counts a
    # shared image repeatedly or, after balancing, still places two distinct
    # near-limit images in one archive. Build consecutive windows against the
    # exact union of resources that the resulting physical EPUB will carry.
    ranges: list[tuple[int, int]] = []
    start = 0
    document_bytes = 0
    resource_names: set[str] = set()
    for index, (document, references) in enumerate(
        zip(documents, referenced_resources, strict=True)
    ):
        next_resources = resource_names | references
        next_document_bytes = document_bytes + len(document.get_content())
        # The fixed allowance covers EPUB navigation, package XML, ZIP
        # headers and metadata. Text XML normally compresses substantially,
        # while image resources are already compressed, so this is a safe,
        # deliberately conservative estimate.
        estimate = (
            next_document_bytes
            + sum(resource_sizes.get(name, 0) for name in next_resources)
            + 128 * 1024
        )
        if index > start and estimate > target:
            ranges.append((start, index))
            start = index
            document_bytes = len(document.get_content())
            resource_names = set(references)
        else:
            document_bytes = next_document_bytes
            resource_names = next_resources
    ranges.append((start, len(documents)))

    # The guard may request more parts for another independently measured
    # workload limit. Split the widest logical ranges without changing order.
    while len(ranges) < min(minimum_parts, len(documents)):
        split_at = max(
            range(len(ranges)),
            key=lambda index: ranges[index][1] - ranges[index][0],
        )
        range_start, range_end = ranges[split_at]
        if range_end - range_start <= 1:
            break
        midpoint = range_start + (range_end - range_start) // 2
        ranges[split_at : split_at + 1] = [
            (range_start, midpoint),
            (midpoint, range_end),
        ]

    parts: list[Part] = []
    identifiers = book.get_metadata("DC", "identifier")
    source_identifier = (
        str(identifiers[0][0]).strip()
        if identifiers and identifiers[0] and identifiers[0][0]
        else source.stem
    )
    index = 0
    while index < len(ranges):
        start, end = ranges[index]
        part_book = epub.EpubBook()
        part_book.set_identifier(f"{source_identifier}-part-{index + 1}")
        titles = book.get_metadata("DC", "title")
        part_book.set_title(str(titles[0][0]) if titles else source.stem)
        for language, _attrs in book.get_metadata("DC", "language"):
            part_book.set_language(str(language))

        # Copy only images actually referenced by this consecutive spine
        # window. Copying the entire resource manifest into every physical
        # part both multiplies storage and can leave every part above 5 MB.
        selected_resource_names: set[str] = set()
        for names in referenced_resources[start:end]:
            selected_resource_names.update(names)
        for resource_name in sorted(selected_resource_names):
            resource = image_resources.get(resource_name)
            if resource is None:
                continue
            try:
                part_book.add_item(copy.deepcopy(resource))
            except Exception:
                pass
        selected = []
        for document in documents[start:end]:
            cloned = epub.EpubHtml(
                uid=document.get_id(),
                file_name=document.file_name,
                media_type=document.media_type,
                title=getattr(document, "title", ""),
                lang=getattr(document, "lang", None),
            )
            cloned.set_content(document.get_content())
            part_book.add_item(cloned)
            selected.append(cloned)
        part_book.spine = ["nav", *selected]
        part_book.toc = tuple(selected)
        part_book.add_item(epub.EpubNcx())
        part_book.add_item(epub.EpubNav())
        path = output / f"epub-{index:06d}.epub"
        epub.write_epub(str(path), part_book)
        part_size = path.stat().st_size
        if part_size > target and end - start > 1:
            # Estimates intentionally err on the safe side, but EPUB writers
            # can add format-specific overhead. Refine the window using the
            # measured output until it is near the configured 70–80% target.
            path.unlink(missing_ok=True)
            midpoint = start + (end - start) // 2
            ranges[index : index + 1] = [(start, midpoint), (midpoint, end)]
            continue
        if part_size > _HARD_BYTES["epub"]:
            raise SplitFailure(
                "atomic_epub_spine_item_too_large",
                (
                    f"EPUB spine window {start}-{end - 1} remains "
                    f"{part_size} bytes after reference-aware splitting"
                ),
            )
        parts.append(
            Part(
                path,
                "epub",
                {
                    "kind": "spine_items",
                    "spine_start": start,
                    "spine_end": end - 1,
                },
                {
                    "spine_items": end - start,
                    "referenced_resources": len(selected_resource_names),
                    "file_bytes": part_size,
                },
            )
        )
        index += 1
    return parts


def _epub_referenced_resources(
    document: Any, image_resources: dict[str, Any]
) -> set[str]:
    from bs4 import BeautifulSoup

    aliases: dict[str, str] = {}
    for name in image_resources:
        normalized = posixpath.normpath(unquote(name))
        aliases[name] = name
        aliases[normalized] = name
        aliases[posixpath.basename(normalized)] = name

    try:
        soup = BeautifulSoup(document.get_content(), "xml")
    except Exception:
        return set()
    document_name = str(getattr(document, "file_name", "") or "").strip()
    base = (
        posixpath.dirname(posixpath.normpath(document_name))
        if document_name
        else ""
    )
    references: set[str] = set()
    for tag in soup.find_all(True):
        candidates: list[str] = []
        for attribute in ("src", "href", "xlink:href", "poster", "data-src"):
            value = str(tag.get(attribute) or "").strip()
            if value:
                candidates.append(value)
        srcset = str(tag.get("srcset") or "")
        for item in srcset.split(","):
            candidate = item.strip().split(" ", 1)[0]
            if candidate:
                candidates.append(candidate)
        for candidate in candidates:
            parsed = urlsplit(unquote(candidate))
            if parsed.scheme in {"http", "https", "data"}:
                continue
            raw_path = parsed.path
            if not raw_path:
                continue
            normalized = posixpath.normpath(posixpath.join(base, raw_path))
            for alias in (
                candidate,
                raw_path,
                normalized,
                posixpath.basename(normalized),
            ):
                resource_name = aliases.get(alias)
                if resource_name is not None:
                    references.add(resource_name)
                    break
    return references


def _split_mhtml(
    source: Path, output: Path, minimum_parts: int, ratio: float
) -> list[Part]:
    from bs4 import BeautifulSoup

    with source.open("rb") as stream:
        message = BytesParser(policy=email_policy.default).parse(stream)
    html_parts = [
        part
        for part in message.walk()
        if part.get_content_type() in {"text/html", "text/plain"}
        and part.get_content_disposition() != "attachment"
    ]
    if not html_parts:
        raise SplitFailure("empty_document", "MHTML contains no readable body")
    html = "\n".join(part.get_content() for part in html_parts)
    soup = BeautifulSoup(html, "lxml")
    raw_units = [
        str(unit)
        for unit in (soup.body or soup).children
        if str(unit).strip()
    ]
    if not raw_units:
        raw_units = [html]
    target = min(
        int(_HARD_BYTES["mhtml"] * ratio),
        max(64 * 1024, math.ceil(source.stat().st_size / minimum_parts)),
    )
    # EmailMessage normally base64-encodes CJK-heavy HTML. Budget for that
    # expansion before grouping, then verify the actual serialized archive.
    html_budget = max(32 * 1024, int(max(1, target - 64 * 1024) / 1.4))
    units: list[_MHTMLUnit] = []
    for source_index, raw_unit in enumerate(raw_units):
        fragments = _mhtml_fragment_dom_unit(raw_unit, html_budget)
        units.extend(
            _MHTMLUnit(
                html=fragment,
                source_index=source_index,
                segment_index=segment_index,
                segment_count=len(fragments),
            )
            for segment_index, fragment in enumerate(fragments)
        )

    resources = [
        part
        for part in message.walk()
        if part.get_content_type().startswith("image/")
    ]
    resource_metadata = [
        (
            resource,
            _mhtml_resource_aliases(resource),
            len(resource.get_payload(decode=True) or b""),
        )
        for resource in resources
    ]

    # Embedded resources are emitted once, immediately after their first
    # logical reference. Re-attaching a shared logo/diagram to every DOM
    # window multiplies storage and produces duplicate multimodal chunks.
    groups: list[list[_MHTMLUnit]] = []
    current: list[_MHTMLUnit] = []
    current_bytes = 0
    for unit in units:
        unit_bytes = len(unit.html.encode("utf-8"))
        if current and current_bytes + unit_bytes > html_budget:
            groups.append(current)
            current = [unit]
            current_bytes = unit_bytes
        else:
            current.append(unit)
            current_bytes += unit_bytes
    if current:
        groups.append(current)

    result: list[Part] = []
    emitted_resources: set[int] = set()
    pending = list(groups)
    subject = str(message.get("Subject", source.stem))
    while pending:
        group = pending.pop(0)
        group_html = "".join(unit.html for unit in group)
        referenced = _mhtml_referenced_resources(group_html, resource_metadata)
        text_html = _mhtml_rewrite_image_references(group_html)
        path = output / f"mhtml-{len(result):06d}.mhtml"
        _write_mhtml_archive(path, subject, text_html, ())
        if path.stat().st_size > _HARD_BYTES["mhtml"]:
            path.unlink(missing_ok=True)
            if len(group) > 1:
                midpoint = max(1, len(group) // 2)
                pending[0:0] = [group[:midpoint], group[midpoint:]]
                continue
            smaller_budget = max(16 * 1024, html_budget // 2)
            fragments = _mhtml_fragment_dom_unit(group[0].html, smaller_budget)
            if len(fragments) <= 1:
                raise SplitFailure(
                    "atomic_mhtml_dom_unit_too_large",
                    (
                        f"MHTML DOM unit {group[0].source_index} remains above "
                        f"{_HARD_BYTES['mhtml']} bytes after structural splitting"
                    ),
                )
            pending[0:0] = [
                [
                    _MHTMLUnit(
                        html=fragment,
                        source_index=group[0].source_index,
                        segment_index=index,
                        segment_count=len(fragments),
                    )
                ]
                for index, fragment in enumerate(fragments)
            ]
            continue

        unit_start = group[0].source_index
        unit_end = group[-1].source_index
        locator: dict[str, Any] = {
            "kind": "dom_units",
            "unit_start": unit_start,
            "unit_end": unit_end,
        }
        if group[0].segment_count > 1 or group[-1].segment_count > 1:
            locator.update(
                {
                    "segment_start": group[0].segment_index,
                    "segment_end": group[-1].segment_index,
                    "segment_count": max(
                        unit.segment_count for unit in group
                    ),
                }
            )
        result.append(
            Part(
                path,
                "mhtml",
                locator,
                {
                    "dom_units": unit_end - unit_start + 1,
                    "dom_segments": len(group),
                    "file_bytes": path.stat().st_size,
                },
            )
        )

        for resource, aliases, _size in referenced:
            resource_key = id(resource)
            if resource_key in emitted_resources:
                continue
            emitted_resources.add(resource_key)
            resource_index = next(
                index
                for index, candidate in enumerate(resource_metadata)
                if candidate[0] is resource
            )
            context, reference_unit = _mhtml_resource_context(
                group, aliases, resource_index
            )
            result.extend(
                _mhtml_resource_parts(
                    output=output,
                    result_offset=len(result),
                    subject=subject,
                    resource=resource,
                    resource_index=resource_index,
                    source_unit=reference_unit,
                    context=context,
                )
            )
    return result


def _write_mhtml_archive(
    path: Path,
    subject: str,
    body_html: str,
    resources: Sequence[Any],
) -> None:
    out = EmailMessage()
    out["Subject"] = subject
    out.make_related()
    body = EmailMessage()
    body.set_content(
        "<html><body>" + body_html + "</body></html>",
        subtype="html",
        charset="utf-8",
    )
    out.attach(body)
    for resource in resources:
        out.attach(copy.deepcopy(resource))
    with path.open("xb") as sink:
        BytesGenerator(sink, policy=email_policy.default).flatten(out)


def _mhtml_fragment_dom_unit(
    value: str, byte_budget: int, depth: int = 0
) -> list[str]:
    """Split an oversized DOM unit while retaining its structural wrappers."""
    from bs4 import BeautifulSoup, NavigableString, Tag

    if len(value.encode("utf-8")) <= byte_budget:
        return [value]
    if depth > 32:
        raise SplitFailure(
            "mhtml_dom_depth_exceeded",
            "MHTML DOM nesting exceeds the safe structural split depth",
        )

    soup = BeautifulSoup(value, "html.parser")
    roots = [
        node
        for node in soup.contents
        if not isinstance(node, NavigableString) or str(node).strip()
    ]
    if len(roots) != 1 or not isinstance(roots[0], Tag):
        fragments: list[str] = []
        for root in roots:
            root_html = str(root)
            if len(root_html.encode("utf-8")) <= byte_budget:
                fragments.append(root_html)
            elif isinstance(root, Tag):
                fragments.extend(
                    _mhtml_fragment_dom_unit(root_html, byte_budget, depth + 1)
                )
            else:
                fragments.extend(
                    html_lib.escape(fragment)
                    for fragment in _semantic_text_fragments(
                        str(root), byte_budget
                    )
                )
        if fragments:
            return fragments
        return [
            html_lib.escape(fragment)
            for fragment in _semantic_text_fragments(value, byte_budget)
        ]

    root = roots[0]
    empty = copy.deepcopy(root)
    empty.clear()
    wrapper_bytes = len(str(empty).encode("utf-8"))
    inner_budget = byte_budget - wrapper_bytes - 256
    if inner_budget < 1024 or not root.contents:
        text = root.get_text()
        if not text:
            raise SplitFailure(
                "atomic_mhtml_dom_unit_too_large",
                "an atomic MHTML DOM element exceeds the parser hard limit",
            )
        return [
            _mhtml_wrap_tag(root, [html_lib.escape(fragment)])
            for fragment in _semantic_text_fragments(
                text, max(1024, inner_budget)
            )
        ]

    child_fragments: list[str] = []
    for child in root.contents:
        child_html = str(child)
        if not child_html:
            continue
        if len(child_html.encode("utf-8")) <= inner_budget:
            child_fragments.append(child_html)
        elif isinstance(child, Tag):
            child_fragments.extend(
                _mhtml_fragment_dom_unit(
                    child_html, inner_budget, depth + 1
                )
            )
        else:
            child_fragments.extend(
                html_lib.escape(fragment)
                for fragment in _semantic_text_fragments(
                    str(child), inner_budget
                )
            )

    result: list[str] = []
    current: list[str] = []
    for fragment in child_fragments:
        candidate = _mhtml_wrap_tag(root, [*current, fragment])
        if current and len(candidate.encode("utf-8")) > byte_budget:
            result.append(_mhtml_wrap_tag(root, current))
            current = [fragment]
        else:
            current.append(fragment)
    if current:
        result.append(_mhtml_wrap_tag(root, current))
    if any(len(fragment.encode("utf-8")) > byte_budget for fragment in result):
        raise SplitFailure(
            "atomic_mhtml_dom_unit_too_large",
            "MHTML DOM unit remains oversized after recursive splitting",
        )
    return result


def _mhtml_wrap_tag(template: Any, children: Sequence[str]) -> str:
    from bs4 import BeautifulSoup

    clone = copy.deepcopy(template)
    clone.clear()
    for child_html in children:
        fragment = BeautifulSoup(child_html, "html.parser")
        for child in list(fragment.contents):
            clone.append(copy.deepcopy(child))
    return str(clone)


def _mhtml_rewrite_image_references(value: str) -> str:
    from bs4 import BeautifulSoup

    soup = BeautifulSoup(value, "html.parser")
    for image in list(soup.find_all("img")):
        reference = str(image.get("src", "") or "")
        alt = str(image.get("alt", "") or "").strip()
        label = alt or posixpath.basename(urlsplit(reference).path) or "embedded image"
        marker = soup.new_tag("span")
        marker["data-logical-image"] = reference[:512]
        marker.string = f"[嵌入图像：{label}]"
        image.replace_with(marker)
    for source in list(soup.find_all("source")):
        if source.get("src") or source.get("srcset"):
            source.decompose()
    return "".join(str(node) for node in soup.contents)


def _mhtml_resource_context(
    units: Sequence[_MHTMLUnit],
    aliases: set[str],
    resource_index: int,
) -> tuple[str, int]:
    from bs4 import BeautifulSoup

    lowered_aliases = {alias.lower() for alias in aliases}
    for unit in units:
        soup = BeautifulSoup(unit.html, "html.parser")
        for image in soup.find_all("img"):
            reference = unquote(str(image.get("src", "") or "")).lower()
            if reference and any(alias in reference for alias in lowered_aliases):
                heading = image.find_previous(
                    ["h1", "h2", "h3", "h4", "h5", "h6"]
                )
                caption = image.find_parent("figure")
                caption = caption.find("figcaption") if caption else None
                values = [
                    heading.get_text(" ", strip=True) if heading else "",
                    str(image.get("alt", "") or "").strip(),
                    caption.get_text(" ", strip=True) if caption else "",
                ]
                context = " — ".join(value for value in values if value)
                return (
                    (context or f"嵌入图像 {resource_index + 1}")[:500],
                    unit.source_index,
                )
    return f"嵌入图像 {resource_index + 1}", units[0].source_index


def _mhtml_primary_reference(resource: Any, resource_index: int) -> str:
    content_id = str(resource.get("Content-ID", "") or "").strip().strip("<>")
    if content_id:
        return f"cid:{content_id}"
    location = str(resource.get("Content-Location", "") or "").strip()
    if location:
        return location
    return f"cid:split-resource-{resource_index}"


def _mhtml_resource_wrapper(
    context: str,
    reference: str,
    *,
    source_unit: int,
    frame_index: int | None = None,
    frame_count: int | None = None,
) -> str:
    detail = f"源 DOM 单元 {source_unit + 1}"
    if frame_index is not None and frame_count is not None:
        detail += f"，图像帧 {frame_index + 1}/{frame_count}"
    return (
        f"<section data-source-dom-unit='{source_unit}'>"
        f"<h2>{html_lib.escape(context)}</h2>"
        f"<p>{html_lib.escape(detail)}</p>"
        f"<img src='{html_lib.escape(reference, quote=True)}' "
        f"alt='{html_lib.escape(context, quote=True)}'/></section>"
    )


def _mhtml_resource_parts(
    *,
    output: Path,
    result_offset: int,
    subject: str,
    resource: Any,
    resource_index: int,
    source_unit: int,
    context: str,
) -> list[Part]:
    hard = _HARD_BYTES["mhtml"]
    reference = _mhtml_primary_reference(resource, resource_index)
    wrapper = _mhtml_resource_wrapper(
        context, reference, source_unit=source_unit
    )
    path = output / f"mhtml-{result_offset:06d}.mhtml"
    _write_mhtml_archive(path, subject, wrapper, (resource,))
    decoded = resource.get_payload(decode=True) or b""
    if path.stat().st_size <= hard:
        return [
            Part(
                path,
                "mhtml",
                {
                    "kind": "dom_units",
                    "unit_start": source_unit,
                    "unit_end": source_unit,
                    "resource_continuation": True,
                    "resource_index": resource_index,
                    "resource_preserved": True,
                },
                {
                    "dom_units": 0,
                    "embedded_resources": 1,
                    "resource_bytes": len(decoded),
                    "file_bytes": path.stat().st_size,
                },
            )
        ]
    path.unlink(missing_ok=True)
    return _mhtml_tiled_resource_parts(
        output=output,
        result_offset=result_offset,
        subject=subject,
        resource=resource,
        resource_index=resource_index,
        source_unit=source_unit,
        context=context,
    )


def _mhtml_tiled_resource_parts(
    *,
    output: Path,
    result_offset: int,
    subject: str,
    resource: Any,
    resource_index: int,
    source_unit: int,
    context: str,
) -> list[Part]:
    from PIL import Image

    decoded = resource.get_payload(decode=True) or b""
    try:
        image = Image.open(io.BytesIO(decoded))
    except Exception as exc:
        raise SplitFailure(
            "atomic_mhtml_resource_too_large",
            (
                f"embedded resource {resource_index} exceeds the MHTML limit "
                f"and cannot be losslessly tiled: {exc}"
            ),
        ) from exc

    parts: list[Part] = []
    try:
        frame_count = max(1, int(getattr(image, "n_frames", 1)))
        for frame_index in range(frame_count):
            image.seek(frame_index)
            image.load()
            width, height = image.size
            if width <= 0 or height <= 0:
                raise SplitFailure(
                    "invalid_mhtml_resource",
                    f"embedded resource {resource_index} has invalid dimensions",
                )
            base = image.convert("RGBA" if "A" in image.mode else "RGB")
            pending: list[tuple[int, int, int, int]] = [(0, 0, width, height)]
            while pending:
                left, top, right, bottom = pending.pop(0)
                tile = base.crop((left, top, right, bottom))
                tile_buffer = io.BytesIO()
                tile.save(tile_buffer, format="PNG", optimize=True)
                tile_bytes = tile_buffer.getvalue()
                tile_index = len(parts)
                content_id = (
                    f"split-resource-{resource_index}-frame-{frame_index}-"
                    f"tile-{tile_index}"
                )
                tile_resource = EmailMessage()
                tile_resource.set_content(
                    tile_bytes, maintype="image", subtype="png"
                )
                tile_resource["Content-ID"] = f"<{content_id}>"
                tile_resource["Content-Location"] = f"images/{content_id}.png"
                wrapper = _mhtml_resource_wrapper(
                    context,
                    f"cid:{content_id}",
                    source_unit=source_unit,
                    frame_index=frame_index,
                    frame_count=frame_count,
                )
                path = output / f"mhtml-{result_offset + len(parts):06d}.mhtml"
                _write_mhtml_archive(path, subject, wrapper, (tile_resource,))
                if path.stat().st_size > _HARD_BYTES["mhtml"]:
                    path.unlink(missing_ok=True)
                    rect_width = right - left
                    rect_height = bottom - top
                    if max(rect_width, rect_height) <= 64:
                        raise SplitFailure(
                            "atomic_mhtml_resource_too_large",
                            (
                                f"embedded resource {resource_index} tile "
                                f"{left},{top},{right},{bottom} exceeds the limit"
                            ),
                        )
                    if rect_width >= rect_height:
                        middle = left + rect_width // 2
                        overlap = max(4, int(rect_width * 0.025))
                        children = [
                            (left, top, min(right, middle + overlap), bottom),
                            (max(left, middle - overlap), top, right, bottom),
                        ]
                    else:
                        middle = top + rect_height // 2
                        overlap = max(4, int(rect_height * 0.025))
                        children = [
                            (left, top, right, min(bottom, middle + overlap)),
                            (left, max(top, middle - overlap), right, bottom),
                        ]
                    pending[0:0] = children
                    continue
                parts.append(
                    Part(
                        path,
                        "mhtml",
                        {
                            "kind": "dom_units",
                            "unit_start": source_unit,
                            "unit_end": source_unit,
                            "resource_continuation": True,
                            "resource_index": resource_index,
                            "resource_preserved": False,
                            "frame_index": frame_index + 1,
                            "frame_count": frame_count,
                            "x_start": left,
                            "y_start": top,
                            "x_end": right,
                            "y_end": bottom,
                            "source_width": width,
                            "source_height": height,
                        },
                        {
                            "dom_units": 0,
                            "embedded_resources": 1,
                            "resource_bytes": len(tile_bytes),
                            "file_bytes": path.stat().st_size,
                        },
                    )
                )
    finally:
        image.close()

    tile_count = len(parts)
    for tile_index, part in enumerate(parts):
        part.locator["resource_tile_index"] = tile_index
        part.locator["resource_tile_count"] = tile_count
    return parts


def _mhtml_resource_aliases(part: Any) -> set[str]:
    aliases: set[str] = set()
    for raw in (
        part.get("Content-Location", ""),
        part.get("Content-ID", ""),
        part.get("X-Attachment-Id", ""),
        part.get_filename() or "",
    ):
        value = unquote(str(raw or "").strip())
        if not value:
            continue
        aliases.add(value)
        aliases.add(value.strip("<>"))
        aliases.add("cid:" + value.strip("<>"))
        parsed = urlsplit(value)
        if parsed.path:
            aliases.add(parsed.path)
            aliases.add(posixpath.basename(parsed.path))
    return {alias for alias in aliases if alias}


def _mhtml_referenced_resources(
    html: str,
    resources: Sequence[tuple[Any, set[str], int]],
) -> list[tuple[Any, set[str], int]]:
    unescaped = unquote(html)
    lowered = unescaped.lower()
    referenced: list[tuple[Any, set[str], int]] = []
    for resource in resources:
        _part, aliases, _size = resource
        if any(alias.lower() in lowered for alias in aliases):
            referenced.append(resource)
    return referenced


def _split_image(
    source: Path,
    output: Path,
    ext: str,
    minimum_parts: int,
    ratio: float,
    policy: SplitPolicy,
) -> list[Part]:
    from PIL import Image

    try:
        image = Image.open(source)
    except Exception as exc:
        raise SplitFailure("invalid_image", f"image is malformed: {exc}") from exc
    try:
        frame_count = max(1, int(getattr(image, "n_frames", 1)))
        frame_sizes: list[tuple[int, int]] = []
        for frame_index in range(frame_count):
            image.seek(frame_index)
            width, height = image.size
            if width <= 0 or height <= 0:
                raise SplitFailure("empty_image", "image dimensions are invalid")
            frame_sizes.append((width, height))

        total_pixels = sum(width * height for width, height in frame_sizes)
        desired_total = max(
            minimum_parts,
            frame_count,
            math.ceil(
                source.stat().st_size
                / max(1, int(_HARD_BYTES[ext] * ratio))
            ),
            math.ceil(total_pixels / policy.image_tile_pixels),
        )
        result: list[Part] = []

        for frame_index, (width, height) in enumerate(frame_sizes):
            image.seek(frame_index)
            image.load()
            base = image.convert("RGBA" if image.mode == "RGBA" else "RGB")
            frame_pixels = width * height
            desired = max(
                1,
                math.ceil(desired_total * frame_pixels / max(1, total_pixels)),
                math.ceil(frame_pixels / policy.image_tile_pixels),
            )
            columns = max(
                1, math.ceil(math.sqrt(desired * width / max(1, height)))
            )
            rows = max(1, math.ceil(desired / columns))
            tile_width = math.ceil(width / columns)
            tile_height = math.ceil(height / rows)
            overlap_x = max(8, int(tile_width * 0.05))
            overlap_y = max(8, int(tile_height * 0.05))
            pending: list[tuple[int, int, int, int]] = []
            for row in range(rows):
                for column in range(columns):
                    pending.append(
                        (
                            max(0, column * tile_width - overlap_x),
                            max(0, row * tile_height - overlap_y),
                            min(width, (column + 1) * tile_width + overlap_x),
                            min(height, (row + 1) * tile_height + overlap_y),
                        )
                    )

            while pending:
                left, top, right, bottom = pending.pop(0)
                if left >= right or top >= bottom:
                    continue
                tile = base.crop((left, top, right, bottom))
                path = output / f"image-{len(result):06d}.png"
                tile.save(path, format="PNG", optimize=True)
                if path.stat().st_size > _HARD_BYTES["png"]:
                    path.unlink(missing_ok=True)
                    rect_width, rect_height = right - left, bottom - top
                    if max(rect_width, rect_height) <= 64:
                        raise SplitFailure(
                            "atomic_image_tile_too_large",
                            f"image tile {left},{top},{right},{bottom} exceeds 5MB",
                        )
                    if rect_width >= rect_height:
                        middle = left + rect_width // 2
                        overlap = max(4, int(rect_width * 0.025))
                        children = [
                            (left, top, min(right, middle + overlap), bottom),
                            (max(left, middle - overlap), top, right, bottom),
                        ]
                    else:
                        middle = top + rect_height // 2
                        overlap = max(4, int(rect_height * 0.025))
                        children = [
                            (left, top, right, min(bottom, middle + overlap)),
                            (left, max(top, middle - overlap), right, bottom),
                        ]
                    pending = children + pending
                    continue
                result.append(
                    Part(
                        path,
                        "png",
                        {
                            "kind": "image_tile",
                            "frame_index": frame_index + 1,
                            "frame_count": frame_count,
                            "x_start": left,
                            "y_start": top,
                            "x_end": right,
                            "y_end": bottom,
                            "source_width": width,
                            "source_height": height,
                        },
                        {
                            "pixels": (right - left) * (bottom - top),
                            "frame_index": frame_index + 1,
                        },
                    )
                )
                if len(result) > policy.max_parts:
                    raise SplitFailure(
                        "too_many_parts",
                        f"image requires more than {policy.max_parts} tiles",
                    )
        return result
    finally:
        image.close()


def _split_audio(
    source: Path,
    output: Path,
    ext: str,
    minimum_parts: int,
    ratio: float,
    policy: SplitPolicy,
) -> list[Part]:
    duration = _probe_duration(source, policy.audio_timeout_seconds)
    if duration <= 0:
        raise SplitFailure("invalid_audio", "audio duration could not be determined")
    desired = max(
        minimum_parts,
        math.ceil(source.stat().st_size / max(1, int(_HARD_BYTES[ext] * ratio))),
    )
    segment = max(1.0, duration / desired)
    overlap = min(2.0, segment * 0.05)
    result: list[Part] = []
    start = 0.0
    while start < duration:
        end = min(duration, start + segment + (overlap if start + segment < duration else 0))
        path = output / f"audio-{len(result):06d}.{_output_extension(ext)}"
        command = [
            "ffmpeg",
            "-hide_banner",
            "-loglevel",
            "error",
            "-nostdin",
            "-ss",
            f"{start:.3f}",
            "-i",
            str(source),
            "-t",
            f"{end - start:.3f}",
            "-map",
            "0:a:0",
            "-c",
            "copy",
            "-avoid_negative_ts",
            "make_zero",
            "-y",
            str(path),
        ]
        completed = subprocess.run(
            command,
            capture_output=True,
            timeout=policy.audio_timeout_seconds,
            check=False,
        )
        if completed.returncode != 0 or not path.is_file():
            raise SplitFailure(
                "audio_split_failed",
                completed.stderr.decode("utf-8", errors="replace")[-2000:],
            )
        if path.stat().st_size > _HARD_BYTES[ext]:
            raise SplitFailure(
                "atomic_audio_segment_too_large",
                f"audio segment {start:.3f}-{end:.3f}s exceeds parser limit",
            )
        result.append(
            Part(
                path,
                ext,
                {"kind": "time_range", "start_seconds": start, "end_seconds": end},
                {"duration_seconds": end - start, "overlap_seconds": overlap},
            )
        )
        start += segment
    return result


def _probe_duration(source: Path, timeout: int) -> float:
    completed = subprocess.run(
        [
            "ffprobe",
            "-v",
            "error",
            "-show_entries",
            "format=duration",
            "-of",
            "default=noprint_wrappers=1:nokey=1",
            str(source),
        ],
        capture_output=True,
        timeout=timeout,
        check=False,
        text=True,
    )
    if completed.returncode != 0:
        return 0.0
    try:
        return float(completed.stdout.strip())
    except ValueError:
        return 0.0


def _libreoffice_convert(
    source: Path, output: Path, target: str, timeout: int
) -> Path:
    output.mkdir(parents=True, exist_ok=True)
    staged = output / f"source{source.suffix.lower()}"
    shutil.copyfile(source, staged)
    profile = output / f"lo-profile-{time.time_ns()}"
    command = [
        "soffice",
        "--headless",
        "--nologo",
        "--nodefault",
        "--nolockcheck",
        "--nofirststartwizard",
        f"-env:UserInstallation={profile.as_uri()}",
        "--convert-to",
        target,
        "--outdir",
        str(output),
        str(staged),
    ]
    try:
        completed = subprocess.run(
            command,
            capture_output=True,
            timeout=timeout,
            check=False,
            text=True,
        )
    except subprocess.TimeoutExpired as exc:
        raise SplitFailure(
            "office_conversion_timeout", f"LibreOffice conversion exceeded {timeout}s", True
        ) from exc
    expected = output / f"{staged.stem}.{target}"
    if completed.returncode != 0 or not expected.is_file():
        detail = (completed.stderr or completed.stdout or "conversion failed")[-2000:]
        raise SplitFailure("office_conversion_failed", detail)
    return expected


def _balanced_ranges(weights: Sequence[int], desired: int) -> list[tuple[int, int]]:
    if not weights:
        return []
    desired = max(1, min(desired, len(weights)))
    remaining_weight = sum(max(1, value) for value in weights)
    ranges: list[tuple[int, int]] = []
    start = 0
    for part_index in range(desired):
        parts_left = desired - part_index
        items_left = len(weights) - start
        if parts_left == 1:
            ranges.append((start, len(weights)))
            break
        target = remaining_weight / parts_left
        end = start
        current = 0
        max_end = len(weights) - (parts_left - 1)
        while end < max_end:
            weight = max(1, weights[end])
            if end > start and current + weight > target:
                before = abs(target - current)
                after = abs(target - current - weight)
                if after >= before:
                    break
            current += weight
            end += 1
            if current >= target:
                break
        if end == start:
            end += 1
            current = max(1, weights[start])
        ranges.append((start, end))
        remaining_weight -= current
        start = end
    return ranges


def _group_strings(values: Sequence[str], target_bytes: int) -> list[list[str]]:
    groups: list[list[str]] = []
    current: list[str] = []
    size = 0
    for value in values:
        value_size = len(value.encode("utf-8"))
        if value_size > target_bytes and not current:
            groups.append([value])
            continue
        if current and size + value_size > target_bytes:
            groups.append(current)
            current = []
            size = 0
        current.append(value)
        size += value_size
    if current:
        groups.append(current)
    return groups


def _csv_dialect(source: Path, encoding: str):
    with source.open("r", encoding=encoding, newline="") as stream:
        sample = stream.read(64 * 1024)
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        return csv.excel


def _detect_text_encoding(source: Path) -> str:
    with source.open("rb") as stream:
        sample = stream.read(256 * 1024)
    if sample.startswith(b"\xef\xbb\xbf"):
        return "utf-8-sig"
    if sample.startswith((b"\xff\xfe", b"\xfe\xff")):
        return "utf-16"
    try:
        sample.decode("utf-8", errors="strict")
        return "utf-8"
    except UnicodeDecodeError:
        try:
            sample.decode("gb18030", errors="strict")
            return "gb18030"
        except UnicodeDecodeError as exc:
            raise SplitFailure(
                "unsupported_text_encoding",
                "text is neither UTF-8 nor a supported GB18030 document",
            ) from exc


def _safe_sheet_title(value: str) -> str:
    forbidden = set("[]:*?/\\")
    sanitized = "".join("_" if char in forbidden else char for char in value).strip()
    return (sanitized or "Sheet")[:31]


def _format_column_header_context(
    values: Sequence[Any], starting_column: int
) -> str:
    """Build a bounded schema label carried with every physical table part."""
    fields: list[str] = []
    total_bytes = 0
    for offset, value in enumerate(values):
        if value is None:
            continue
        text = " ".join(str(value).split())
        if not text:
            continue
        if len(text) > 256:
            text = text[:253] + "..."
        field = f"{_xlsx_column_name(starting_column + offset)}={text}"
        encoded_bytes = len(field.encode("utf-8"))
        if total_bytes + encoded_bytes > 12 * 1024:
            break
        fields.append(field)
        total_bytes += encoded_bytes
    return "；".join(fields)


def _xlsx_column_name(column: int) -> str:
    result = ""
    current = max(1, int(column))
    while current:
        current, remainder = divmod(current - 1, 26)
        result = chr(ord("A") + remainder) + result
    return result


def _normalize_ext(value: str) -> str:
    value = str(value or "").strip().lower()
    if "/" in value:
        aliases = {
            "application/pdf": "pdf",
            "application/json": "json",
            "text/plain": "txt",
            "text/csv": "csv",
            "text/markdown": "md",
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document": "docx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": "xlsx",
            "application/vnd.openxmlformats-officedocument.presentationml.presentation": "pptx",
            "application/vnd.ms-powerpoint": "ppt",
        }
        return aliases.get(value, value)
    return value.lstrip(".")


def _output_extension(value: str) -> str:
    if value == "text":
        return "txt"
    if value == "markdown":
        return "md"
    return value


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        while chunk := source.read(_MB):
            digest.update(chunk)
    return digest.hexdigest()


def _positive_int_env(name: str, fallback: int) -> int:
    try:
        value = int(os.getenv(name, ""))
    except ValueError:
        return fallback
    return value if value > 0 else fallback


def _positive_float_env(name: str, fallback: float) -> float:
    try:
        value = float(os.getenv(name, ""))
    except ValueError:
        return fallback
    return value if value > 0 else fallback


def _ratio_env(name: str, fallback: float) -> float:
    value = _positive_float_env(name, fallback)
    return value if 0.5 <= value <= 0.9 else fallback
