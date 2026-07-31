from __future__ import annotations

import argparse
import hashlib
import html
import json
import os
import re
import subprocess
import sys
import time
import unicodedata
import uuid
from collections import Counter
from datetime import date, datetime, time as datetime_time
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterable, Mapping

from openpyxl import load_workbook

from cluster_e2e import APIClient, E2EFailure, generated_questions


TERMINAL_STAGE_STATUSES = {"completed", "done", "degraded", "failed", "skipped"}
FAILED_STAGE_STATUSES = {"failed", "cancelled"}
UUID_RE = re.compile(r"^[0-9a-fA-F-]{36}$")


def emit(event: str, **fields: Any) -> None:
    print(
        json.dumps(
            {"timestamp": datetime.now().astimezone().isoformat(), "event": event, **fields},
            ensure_ascii=False,
            default=str,
        ),
        flush=True,
    )


def require_api_key() -> str:
    value = os.environ.get("WEKNORA_E2E_TENANT_API_KEY", "").strip()
    if not value:
        raise E2EFailure("WEKNORA_E2E_TENANT_API_KEY is required")
    return value


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _psql_json(container: str, sql: str) -> Any:
    command = [
        "docker",
        "exec",
        container,
        "psql",
        "-U",
        "postgres",
        "-d",
        "WeKnora",
        "-X",
        "-qAt",
        "-v",
        "ON_ERROR_STOP=1",
        "-c",
        sql,
    ]
    completed = subprocess.run(
        command,
        check=False,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
    )
    if completed.returncode != 0:
        raise E2EFailure(
            "local PostgreSQL audit failed: " + completed.stderr.strip()[:2000]
        )
    lines = [line.strip() for line in completed.stdout.splitlines() if line.strip()]
    if not lines:
        return []
    try:
        return json.loads(lines[-1])
    except json.JSONDecodeError as exc:
        raise E2EFailure(
            f"local PostgreSQL audit returned invalid JSON: {lines[-1][:1000]}"
        ) from exc


def _uuid_sql_list(values: Iterable[str]) -> str:
    normalized = []
    for value in values:
        value = value.strip()
        if not UUID_RE.fullmatch(value):
            raise E2EFailure(f"unexpected knowledge id: {value!r}")
        normalized.append("'" + value + "'")
    if not normalized:
        raise E2EFailure("knowledge id list is empty")
    return ",".join(normalized)


def audit_database(container: str, knowledge_ids: list[str]) -> dict[str, Any]:
    ids = _uuid_sql_list(knowledge_ids)
    plans = _psql_json(
        container,
        f"""
        SELECT COALESCE(json_agg(row_to_json(q) ORDER BY q.knowledge_id, q.created_at), '[]'::json)
        FROM (
          SELECT id, knowledge_id, processing_generation, planner_version, strategy,
                 state, part_count, completed_parts, failed_parts, total_part_bytes,
                 attempt, last_error, created_at, updated_at
          FROM custom_document_split_plans
          WHERE tenant_id = 10000 AND knowledge_id IN ({ids})
        ) q
        """,
    )
    parts = _psql_json(
        container,
        f"""
        SELECT COALESCE(json_agg(row_to_json(q) ORDER BY q.knowledge_id, q.part_index), '[]'::json)
        FROM (
          SELECT plan_id, knowledge_id, processing_generation, part_index, state,
                 input_size, output_size, markdown_chars, draft_chunks,
                 attempt, last_error, locator, metrics
          FROM custom_document_split_parts
          WHERE tenant_id = 10000 AND knowledge_id IN ({ids})
        ) q
        """,
    )
    work_items = _psql_json(
        container,
        f"""
        SELECT COALESCE(json_agg(row_to_json(q) ORDER BY q.knowledge_id, q.work_kind, q.item_id), '[]'::json)
        FROM (
          SELECT id, knowledge_id, processing_generation, item_id, work_kind, state,
                 provider_attempts, materialize_attempts, finalize_attempts,
                 last_error_class, last_error_code, last_error_message,
                 created_at, completed_at
          FROM custom_derivative_work_items
          WHERE tenant_id = 10000 AND knowledge_id IN ({ids})
        ) q
        """,
    )
    provider_calls = _psql_json(
        container,
        f"""
        SELECT COALESCE(json_agg(row_to_json(q) ORDER BY q.knowledge_id, q.work_kind, q.attempt), '[]'::json)
        FROM (
          SELECT w.knowledge_id, w.work_kind, c.attempt, c.provider_request_id,
                 c.processing_generation, c.disposition, c.validation_error,
                 c.response_size, c.created_at, c.validated_at
          FROM custom_derivative_provider_calls c
          JOIN custom_derivative_work_items w ON w.id = c.work_item_id
          WHERE w.tenant_id = 10000 AND w.knowledge_id IN ({ids})
        ) q
        """,
    )
    chunk_rows = _psql_json(
        container,
        f"""
        SELECT COALESCE(json_agg(row_to_json(q) ORDER BY q.knowledge_id), '[]'::json)
        FROM (
          SELECT knowledge_id,
                 count(*) FILTER (WHERE deleted_at IS NULL) AS total_chunks,
                 count(*) FILTER (WHERE deleted_at IS NULL AND is_enabled) AS enabled_chunks,
                 count(*) FILTER (WHERE deleted_at IS NULL AND processing_generation <> '') AS split_chunks,
                 count(DISTINCT split_part_index) FILTER (
                   WHERE deleted_at IS NULL AND processing_generation <> ''
                 ) AS split_part_indexes,
                 min(split_part_index) FILTER (
                   WHERE deleted_at IS NULL AND processing_generation <> ''
                 ) AS min_split_part_index,
                 max(split_part_index) FILTER (
                   WHERE deleted_at IS NULL AND processing_generation <> ''
                 ) AS max_split_part_index
          FROM chunks
          WHERE tenant_id = 10000 AND knowledge_id IN ({ids})
          GROUP BY knowledge_id
        ) q
        """,
    )
    return {
        "plans": plans,
        "parts": parts,
        "work_items": work_items,
        "provider_calls": provider_calls,
        "chunk_rows": chunk_rows,
    }


def validate_split_rows(database: Mapping[str, Any], knowledge_ids: list[str]) -> dict[str, Any]:
    plans = [row for row in database.get("plans", []) if isinstance(row, Mapping)]
    parts = [row for row in database.get("parts", []) if isinstance(row, Mapping)]
    summaries: dict[str, Any] = {}
    for knowledge_id in knowledge_ids:
        scoped_plans = [row for row in plans if row.get("knowledge_id") == knowledge_id]
        if not scoped_plans:
            summaries[knowledge_id] = {
                "mode": "original_file_no_physical_split",
                "plan_count": 0,
                "complete": True,
            }
            continue
        current = scoped_plans[-1]
        plan_parts = [row for row in parts if row.get("plan_id") == current.get("id")]
        indexes = sorted(int(row.get("part_index", -1)) for row in plan_parts)
        expected_count = int(current.get("part_count", 0))
        contiguous = indexes == list(range(expected_count))
        completed = all(str(row.get("state", "")) == "completed" for row in plan_parts)
        complete = (
            str(current.get("state", "")) == "completed"
            and expected_count == len(plan_parts)
            and int(current.get("completed_parts", 0)) == expected_count
            and int(current.get("failed_parts", 0)) == 0
            and contiguous
            and completed
        )
        summaries[knowledge_id] = {
            "mode": "physical_split",
            "plan_count": len(scoped_plans),
            "plan_state": current.get("state"),
            "part_count": expected_count,
            "persisted_parts": len(plan_parts),
            "completed_parts": current.get("completed_parts"),
            "failed_parts": current.get("failed_parts"),
            "indexes_contiguous": contiguous,
            "all_parts_completed": completed,
            "complete": complete,
        }
        if not complete:
            raise E2EFailure(
                f"incomplete physical split for {knowledge_id}: {summaries[knowledge_id]}"
            )
    return summaries


def canonical(value: str) -> str:
    # The chunk API deliberately HTML-escapes quotes and other punctuation for
    # safe display. Decode that transport representation before comparing it
    # with source workbook values; the stored chunk remains unchanged.
    value = html.unescape(value)
    value = unicodedata.normalize("NFKC", value).casefold()
    value = value.replace("\u00a0", "")
    return "".join(character for character in value if character.isalnum())


def cell_variants(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, bool):
        return ["true" if value else "false", "是" if value else "否"]
    if isinstance(value, datetime):
        return [
            value.isoformat(sep=" "),
            value.strftime("%Y-%m-%d %H:%M:%S"),
            value.strftime("%Y/%m/%d %H:%M:%S"),
            value.strftime("%Y年%m月%d日"),
        ]
    if isinstance(value, date):
        return [
            value.isoformat(),
            value.strftime("%Y/%m/%d"),
            value.strftime("%Y年%m月%d日"),
        ]
    if isinstance(value, datetime_time):
        return [value.isoformat(), value.strftime("%H:%M:%S")]
    if isinstance(value, float):
        variants = [str(value), format(value, ".17g"), format(value, ".15g")]
        variants.extend(
            format(value, f".{digits}f").rstrip("0").rstrip(".")
            for digits in range(0, 16)
        )
        if value.is_integer():
            variants.append(str(int(value)))
        return list(dict.fromkeys(variants))
    if isinstance(value, int):
        return [str(value)]
    text = str(value).strip()
    if not text:
        return []
    variants = [text]
    if text.startswith("="):
        variants.append(text[1:])
    try:
        number = Decimal(text.replace(",", ""))
    except InvalidOperation:
        return variants
    variants.append(format(number.normalize(), "f"))
    return list(dict.fromkeys(variants))


def _semantic_cells(path: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    value_book = load_workbook(path, read_only=False, data_only=True, keep_links=False)
    formula_book = load_workbook(path, read_only=False, data_only=False, keep_links=False)
    cells: list[dict[str, Any]] = []
    formulas = 0
    uncached_formulas = 0
    merged_ranges = 0
    try:
        for sheet_index, value_sheet in enumerate(value_book.worksheets):
            formula_sheet = formula_book.worksheets[sheet_index]
            merged_ranges += len(value_sheet.merged_cells.ranges)
            for row in formula_sheet.iter_rows():
                for formula_cell in row:
                    formula = formula_cell.value
                    cached = value_sheet.cell(
                        row=formula_cell.row, column=formula_cell.column
                    ).value
                    if isinstance(formula, str) and formula.startswith("="):
                        formulas += 1
                        if cached is None:
                            uncached_formulas += 1
                    value = cached
                    if value is None and isinstance(formula, str) and formula.startswith("="):
                        value = formula
                    if value is None:
                        continue
                    variants = [canonical(item) for item in cell_variants(value)]
                    variants = [item for item in dict.fromkeys(variants) if item]
                    if not variants:
                        continue
                    cells.append(
                        {
                            "sheet": value_sheet.title,
                            "coordinate": formula_cell.coordinate,
                            "value": str(value)[:300],
                            "variants": variants,
                        }
                    )
        return cells, {
            "sheet_count": len(value_book.worksheets),
            "nonempty_cells": len(cells),
            "formula_cells": formulas,
            "formula_without_cached_value": uncached_formulas,
            "merged_ranges": merged_ranges,
        }
    finally:
        value_book.close()
        formula_book.close()


def audit_chunk_coverage(path: Path, chunks: list[Mapping[str, Any]]) -> dict[str, Any]:
    ordered = sorted(
        chunks,
        key=lambda item: (
            int(item.get("split_part_index", -1) or -1),
            int(item.get("chunk_index", 0) or 0),
            str(item.get("id", "")),
        ),
    )
    text = "\n".join(str(item.get("content", "")) for item in ordered)
    haystack = canonical(text)
    cells, inventory = _semantic_cells(path)
    missing: list[dict[str, Any]] = []
    covered = 0
    auditable = 0
    skipped_short = 0
    for cell in cells:
        variants = cell["variants"]
        # Single-character labels and numbers are too collision-prone to prove
        # independently. They remain covered by the lossless part projection
        # audit; the API content audit reports them separately.
        if max(len(value) for value in variants) < 2:
            skipped_short += 1
            continue
        auditable += 1
        if any(value in haystack for value in variants):
            covered += 1
            continue
        # A long cell can span downstream RAG chunks. Chunk overlap inserts a
        # repeated window between adjacent chunks, so the canonical whole cell
        # is no longer one literal substring even though every source segment
        # is present in order. Verify ordered 64-character segments as the
        # lossless alternative for long values.
        longest = max(variants, key=len)
        if len(longest) >= 128:
            position = 0
            ordered = True
            for offset in range(0, len(longest), 64):
                segment = longest[offset : offset + 64]
                found = haystack.find(segment, position)
                if found < 0:
                    ordered = False
                    break
                position = found + len(segment)
            if ordered:
                covered += 1
                continue
        missing.append(
            {
                "sheet": cell["sheet"],
                "coordinate": cell["coordinate"],
                "value": cell["value"],
            }
        )
    ratio = covered / auditable if auditable else 1.0
    return {
        **inventory,
        "chunk_count": len(chunks),
        "chunk_characters": len(text),
        "auditable_cells": auditable,
        "covered_cells": covered,
        "skipped_single_character_cells": skipped_short,
        "coverage_ratio": ratio,
        "missing_cells": missing[:100],
        "missing_cell_count": len(missing),
        "complete": len(missing) == 0,
    }


def terminal(item: Mapping[str, Any]) -> bool:
    parse_status = str(item.get("parse_status", "")).strip().lower()
    if parse_status in FAILED_STAGE_STATUSES:
        return True
    if parse_status != "completed":
        return False
    enrichment = str(item.get("enrichment_status", "")).strip().lower()
    wiki = str(item.get("wiki_status", "")).strip().lower()
    try:
        pending = int(item.get("pending_subtasks_count", 0) or 0)
    except (TypeError, ValueError):
        return False
    return (
        enrichment in TERMINAL_STAGE_STATUSES
        and wiki in TERMINAL_STAGE_STATUSES
        and pending == 0
    )


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Upload real XLSX files to local WeKnora and prove split/content completeness."
    )
    parser.add_argument("--base-url", default="http://127.0.0.1:8080")
    parser.add_argument("--kb-id", required=True)
    parser.add_argument("--file", action="append", type=Path, required=True)
    parser.add_argument(
        "--reparse-existing",
        action="store_true",
        help="reparse the latest matching knowledge row instead of uploading a duplicate",
    )
    parser.add_argument(
        "--audit-existing",
        action="store_true",
        help="wait for and audit the latest matching knowledge rows without mutating them",
    )
    parser.add_argument("--postgres-container", default="WeKnora-postgres-dev")
    parser.add_argument("--timeout", type=float, default=2400.0)
    parser.add_argument("--poll-interval", type=float, default=3.0)
    parser.add_argument("--report", type=Path, required=True)
    args = parser.parse_args()
    if args.reparse_existing and args.audit_existing:
        raise E2EFailure("--reparse-existing and --audit-existing are mutually exclusive")

    files = [path.resolve() for path in args.file]
    for path in files:
        if not path.is_file() or path.suffix.lower() != ".xlsx":
            raise E2EFailure(f"XLSX file does not exist: {path}")

    client = APIClient(args.base_url, require_api_key(), timeout=120.0)
    kb = client.get_knowledge_base(args.kb_id)
    run_id = f"prod-xlsx-local-{datetime.now().strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:6]}"
    uploads: list[dict[str, Any]] = []
    if args.reparse_existing or args.audit_existing:
        existing = client.list_all_knowledge(args.kb_id)
        selected_ids: set[str] = set()
        for path in files:
            content_sha = sha256_file(path)
            candidates = [
                item
                for item in existing
                if str(item.get("file_name") or item.get("title") or "") == path.name
                and str(item.get("id", "")) not in selected_ids
            ]
            if not candidates:
                raise E2EFailure(f"no existing knowledge row matches {path.name!r}")
            candidates.sort(key=lambda item: str(item.get("created_at", "")))
            selected = candidates[-1]
            knowledge_id = str(selected.get("id", ""))
            if not UUID_RE.fullmatch(knowledge_id):
                raise E2EFailure(f"matching knowledge row has invalid id: {selected!r}")
            selected_ids.add(knowledge_id)
            uploads.append(
                {
                    "knowledge_id": knowledge_id,
                    "file": str(path),
                    "file_name": path.name,
                    "size_bytes": path.stat().st_size,
                    "sha256": content_sha,
                    "previous_processing_generation": str(
                        selected.get("processing_generation", "")
                    ),
                }
            )
        if args.reparse_existing:
            client.request(
                "POST",
                "/api/v1/knowledge/batch-reparse",
                json_body={
                    "kb_id": args.kb_id,
                    "ids": [row["knowledge_id"] for row in uploads],
                },
                timeout=120.0,
            )
            emit(
                "reparse.submitted",
                knowledge_ids=[row["knowledge_id"] for row in uploads],
            )
    else:
        for path in files:
            content = path.read_bytes()
            uploaded = client.upload_document(
                args.kb_id,
                path.name,
                content,
                metadata={
                    "local_e2e_run_id": run_id,
                    "source_sha256": hashlib.sha256(content).hexdigest(),
                },
            )
            row = {
                "knowledge_id": str(uploaded["id"]),
                "file": str(path),
                "file_name": path.name,
                "size_bytes": len(content),
                "sha256": hashlib.sha256(content).hexdigest(),
            }
            uploads.append(row)
            emit("upload.completed", **row)

    deadline = time.monotonic() + args.timeout
    last_snapshot: dict[str, tuple[str, str, str, int, str]] = {}
    final_items: dict[str, Mapping[str, Any]] = {}
    while time.monotonic() < deadline:
        for upload in uploads:
            knowledge_id = upload["knowledge_id"]
            item = client.get_knowledge(knowledge_id)
            snapshot = (
                str(item.get("parse_status", "")),
                str(item.get("enrichment_status", "")),
                str(item.get("wiki_status", "")),
                int(item.get("pending_subtasks_count", 0) or 0),
                str(item.get("processing_generation", "")),
            )
            if last_snapshot.get(knowledge_id) != snapshot:
                emit(
                    "document.progress",
                    knowledge_id=knowledge_id,
                    file_name=upload["file_name"],
                    parse_status=snapshot[0],
                    enrichment_status=snapshot[1],
                    wiki_status=snapshot[2],
                    pending_subtasks_count=snapshot[3],
                    processing_generation=snapshot[4],
                )
                last_snapshot[knowledge_id] = snapshot
            if terminal(item):
                if args.reparse_existing and str(item.get("processing_generation", "")) == str(
                    upload.get("previous_processing_generation", "")
                ):
                    # Batch reparse is asynchronous. The old generation may
                    # still be terminal for a few milliseconds after the API
                    # accepts the request; never audit stale chunks as the new
                    # result.
                    continue
                final_items[knowledge_id] = item
        if len(final_items) == len(uploads):
            break
        time.sleep(args.poll_interval)
    if len(final_items) != len(uploads):
        pending = [
            {"knowledge_id": row["knowledge_id"], "file_name": row["file_name"], "last": last_snapshot.get(row["knowledge_id"])}
            for row in uploads
            if row["knowledge_id"] not in final_items
        ]
        raise E2EFailure(f"documents did not reach terminal state: {pending}")

    knowledge_ids = [row["knowledge_id"] for row in uploads]
    database = audit_database(args.postgres_container, knowledge_ids)
    split_audit = validate_split_rows(database, knowledge_ids)
    all_chunk_types = [
        "text",
        "parent_text",
        "summary",
        "image_ocr",
        "image_caption",
        "table_summary",
        "table_column",
        "entity",
        "relationship",
    ]
    reports = []
    failures = []
    for upload in uploads:
        knowledge_id = upload["knowledge_id"]
        item = final_items[knowledge_id]
        chunks = client.list_chunks(knowledge_id, all_chunk_types)
        text_chunks = [chunk for chunk in chunks if str(chunk.get("chunk_type", "text")) == "text"]
        searchable_chunks = [
            chunk
            for chunk in chunks
            if str(chunk.get("chunk_type", "text")) in {"text", "parent_text"}
        ]
        # Parent/child chunking may split one multiline Excel cell across two
        # embedded child windows. Retrieval expands a hit back to parent_text,
        # so end-to-end information coverage must audit the same searchable
        # context projection while generated questions remain child-only.
        coverage = audit_chunk_coverage(Path(upload["file"]), searchable_chunks)
        question_count = sum(len(generated_questions(chunk)) for chunk in text_chunks)
        result = {
            **upload,
            "parse_status": item.get("parse_status"),
            "core_status": item.get("core_status"),
            "summary_status": item.get("summary_status"),
            "enrichment_status": item.get("enrichment_status"),
            "enrichment_error_summary": item.get("enrichment_error_summary"),
            "wiki_status": item.get("wiki_status"),
            "wiki_error_message": item.get("wiki_error_message"),
            "pending_subtasks_count": item.get("pending_subtasks_count"),
            "split": split_audit[knowledge_id],
            "content_coverage": coverage,
            "generated_question_count": question_count,
            "chunk_type_counts": dict(
                Counter(str(chunk.get("chunk_type", "text")) for chunk in chunks)
            ),
        }
        reports.append(result)
        if str(item.get("parse_status", "")).lower() != "completed":
            failures.append(f"{upload['file_name']}: parse={item.get('parse_status')}")
        if not coverage["complete"]:
            failures.append(
                f"{upload['file_name']}: {coverage['missing_cell_count']} semantic cells missing from chunks"
            )
        emit(
            "document.audited",
            knowledge_id=knowledge_id,
            file_name=upload["file_name"],
            parse_status=item.get("parse_status"),
            enrichment_status=item.get("enrichment_status"),
            wiki_status=item.get("wiki_status"),
            split_complete=split_audit[knowledge_id]["complete"],
            chunk_coverage=coverage["coverage_ratio"],
            missing_cells=coverage["missing_cell_count"],
            generated_questions=question_count,
        )

    report = {
        "run_id": run_id,
        "base_url": args.base_url,
        "knowledge_base": {
            "id": args.kb_id,
            "name": kb.get("name"),
            "question_generation_config": kb.get("question_generation_config"),
            "wiki_config": kb.get("wiki_config"),
            "indexing_strategy": kb.get("indexing_strategy"),
        },
        "documents": reports,
        "database": database,
        "failures": failures,
        "passed": not failures,
    }
    args.report.parent.mkdir(parents=True, exist_ok=True)
    args.report.write_text(
        json.dumps(report, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )
    emit(
        "run.completed",
        run_id=run_id,
        document_count=len(reports),
        passed=not failures,
        report=str(args.report.resolve()),
    )
    if failures:
        raise E2EFailure("; ".join(failures))
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except E2EFailure as exc:
        print(str(exc), file=sys.stderr)
        raise SystemExit(1)
