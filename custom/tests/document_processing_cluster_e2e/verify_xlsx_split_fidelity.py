from __future__ import annotations

import argparse
import json
import math
import tempfile
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from weknora_document_splitter import service
from weknora_document_splitter.xlsx_semantic import inspect_xlsx_semantic_ranges


def _normal(value: Any) -> str:
    if value is None:
        return "<NULL>"
    return f"{type(value).__name__}:{value}"


def _equivalent(expected: Any, actual: Any) -> bool:
    if isinstance(expected, float) and isinstance(actual, float):
        return math.isclose(expected, actual, rel_tol=1e-12, abs_tol=1e-12)
    return type(expected) is type(actual) and expected == actual


def _merged_values(sheet, formula_sheet, row: int, columns: list[int]) -> tuple[Any, ...]:
    values = {column: sheet.cell(row=row, column=column).value for column in columns}
    for column in columns:
        formula = formula_sheet.cell(row=row, column=column).value
        if (
            values[column] is None
            and isinstance(formula, str)
            and formula.startswith("=")
        ):
            values[column] = formula
    for merged in sheet.merged_cells.ranges:
        if row < merged.min_row or row > merged.max_row:
            continue
        top_left = sheet.cell(row=merged.min_row, column=merged.min_col).value
        if top_left is None:
            continue
        for column in columns:
            if merged.min_col <= column <= merged.max_col and values[column] is None:
                values[column] = top_left
    return tuple(values[column] for column in columns)


def audit_workbook(source: Path) -> dict[str, Any]:
    bounds = inspect_xlsx_semantic_ranges(source)
    value_book = load_workbook(source, read_only=False, data_only=True, keep_links=False)
    formula_book = load_workbook(source, read_only=False, data_only=False, keep_links=False)
    formula_cells = 0
    formula_without_cache = 0
    try:
        for sheet_index, bound in enumerate(bounds):
            values = value_book.worksheets[sheet_index]
            formulas = formula_book.worksheets[sheet_index]
            for row in range(1, bound.semantic_max_row + 1):
                for column in range(1, bound.semantic_max_column + 1):
                    formula = formulas.cell(row=row, column=column).value
                    if isinstance(formula, str) and formula.startswith("="):
                        formula_cells += 1
                        if values.cell(row=row, column=column).value is None:
                            formula_without_cache += 1

        with tempfile.TemporaryDirectory() as temp:
            parts = service._split_xlsx(
                source,
                Path(temp),
                minimum_parts=219,
                ratio=0.75,
                policy=service.SplitPolicy(max_parts=10_000),
            )
            grouped: dict[tuple[str, int, int], list[service.Part]] = defaultdict(list)
            for part in parts:
                grouped[
                    (
                        str(part.locator["sheet"]),
                        int(part.locator["column_start"]),
                        int(part.locator["column_end"]),
                    )
                ].append(part)

            mismatches: list[dict[str, Any]] = []
            numeric_roundtrip_cells = 0
            maximum_numeric_relative_error = 0.0
            bit_exact = True
            compared_rows = 0
            compared_nonempty_cells = 0
            sheet_by_name = {sheet.title: sheet for sheet in value_book.worksheets}
            formula_sheet_by_name = {
                sheet.title: sheet for sheet in formula_book.worksheets
            }
            for (sheet_name, column_start, column_end), window_parts in grouped.items():
                source_sheet = sheet_by_name[sheet_name]
                formula_sheet = formula_sheet_by_name[sheet_name]
                for part in sorted(window_parts, key=lambda item: int(item.locator["row_start"])):
                    anchors = [int(column) for column in part.locator["anchor_columns"]]
                    columns = anchors + list(range(column_start, column_end + 1))
                    expected: list[tuple[Any, ...]] = []
                    for row in range(
                        int(part.locator["row_start"]),
                        int(part.locator["row_end"]) + 1,
                    ):
                        projected = _merged_values(
                            source_sheet, formula_sheet, row, columns
                        )
                        if any(value is not None for value in projected):
                            expected.append(projected)

                    parsed = load_workbook(
                        part.path, read_only=True, data_only=False, keep_links=False
                    )
                    try:
                        actual = list(parsed.active.iter_rows(values_only=True))
                    finally:
                        parsed.close()
                    if part.locator["header_repeated"] and actual:
                        actual = actual[1:]
                    width = len(columns)
                    actual = [tuple(row) + (None,) * (width - len(row)) for row in actual]
                    actual = [tuple(row[:width]) for row in actual]
                    compared_rows += len(expected)
                    compared_nonempty_cells += sum(
                        value is not None for row in expected for value in row
                    )
                    normal_actual = [tuple(map(_normal, row)) for row in actual]
                    normal_expected = [tuple(map(_normal, row)) for row in expected]
                    if normal_actual != normal_expected:
                        bit_exact = False
                    for expected_row, actual_row in zip(expected, actual):
                        for expected_value, actual_value in zip(expected_row, actual_row):
                            if (
                                isinstance(expected_value, float)
                                and isinstance(actual_value, float)
                                and expected_value != actual_value
                                and _equivalent(expected_value, actual_value)
                            ):
                                numeric_roundtrip_cells += 1
                                denominator = max(abs(expected_value), 1.0)
                                maximum_numeric_relative_error = max(
                                    maximum_numeric_relative_error,
                                    abs(expected_value - actual_value) / denominator,
                                )
                    semantic_equal = len(actual) == len(expected) and all(
                        _equivalent(expected_value, actual_value)
                        for expected_row, actual_row in zip(expected, actual)
                        for expected_value, actual_value in zip(expected_row, actual_row)
                    )
                    if not semantic_equal:
                        first_difference = None
                        for offset in range(max(len(normal_actual), len(normal_expected))):
                            expected_row = (
                                normal_expected[offset]
                                if offset < len(normal_expected)
                                else None
                            )
                            actual_row = (
                                normal_actual[offset]
                                if offset < len(normal_actual)
                                else None
                            )
                            if expected_row != actual_row:
                                first_difference = {
                                    "row_offset": offset,
                                    "expected": expected_row,
                                    "actual": actual_row,
                                }
                                break
                        mismatches.append(
                            {
                                "sheet": sheet_name,
                                "row_start": part.locator["row_start"],
                                "row_end": part.locator["row_end"],
                                "column_start": column_start,
                                "column_end": column_end,
                                "expected_rows": len(expected),
                                "actual_rows": len(actual),
                                "first_difference": first_difference,
                            }
                        )

            return {
                "file": source.name,
                "part_count": len(parts),
                "sheet_count": len(bounds),
                "raw_cells": sum(item.raw_cells for item in bounds),
                "semantic_cells": sum(item.semantic_cells for item in bounds),
                "style_only_cells": sum(item.style_only_cells for item in bounds),
                "formula_cells": formula_cells,
                "formula_without_cached_value": formula_without_cache,
                "compared_rows": compared_rows,
                "compared_nonempty_cells": compared_nonempty_cells,
                "mismatches": mismatches,
                "numeric_roundtrip_cells": numeric_roundtrip_cells,
                "maximum_numeric_relative_error": maximum_numeric_relative_error,
                "bit_exact_value_projection": bit_exact,
                "lossless_semantic_projection": not mismatches,
            }
    finally:
        value_book.close()
        formula_book.close()


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("files", nargs="*", type=Path)
    parser.add_argument("--directory", action="append", type=Path, default=[])
    parser.add_argument("--summary", action="store_true")
    args = parser.parse_args()
    paths = list(args.files)
    for directory in args.directory:
        paths.extend(sorted(directory.glob("*.xlsx")))
    paths = list(dict.fromkeys(path.resolve() for path in paths))
    reports = []
    for path in paths:
        try:
            reports.append(audit_workbook(path))
        except Exception as exc:
            reports.append(
                {
                    "file": path.name,
                    "error": str(exc),
                    "lossless_semantic_projection": False,
                }
            )
    if args.summary:
        failed = [
            report
            for report in reports
            if not report["lossless_semantic_projection"]
        ]
        output: Any = {
            "file_count": len(reports),
            "passed": len(reports) - len(failed),
            "failed": len(failed),
            "part_count": sum(report.get("part_count", 0) for report in reports),
            "compared_rows": sum(report.get("compared_rows", 0) for report in reports),
            "compared_nonempty_cells": sum(
                report.get("compared_nonempty_cells", 0) for report in reports
            ),
            "formula_cells": sum(report.get("formula_cells", 0) for report in reports),
            "formula_without_cached_value": sum(
                report.get("formula_without_cached_value", 0) for report in reports
            ),
            "numeric_roundtrip_cells": sum(
                report.get("numeric_roundtrip_cells", 0) for report in reports
            ),
            "maximum_numeric_relative_error": max(
                (
                    report.get("maximum_numeric_relative_error", 0.0)
                    for report in reports
                ),
                default=0.0,
            ),
            "failed_reports": failed,
        }
    else:
        output = reports
    print(json.dumps(output, ensure_ascii=False, indent=2, default=str))
    return 0 if all(report["lossless_semantic_projection"] for report in reports) else 1


if __name__ == "__main__":
    raise SystemExit(main())
