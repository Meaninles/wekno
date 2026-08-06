"""Fill merged cell values before pandas reads an XLSX workbook."""

from __future__ import annotations

import logging
import zipfile
from io import BytesIO

logger = logging.getLogger(__name__)


def fill_merged_cells_xlsx(content: bytes) -> bytes:
    """Unmerge ranges and copy the master cell value into every covered cell.

    openpyxl only stores values on the top-left cell of a merge; pandas then
    sees NaN in the rest. Filling makes row-wise RAG chunks retain context.
    """
    if not zipfile.is_zipfile(BytesIO(content)):
        return content

    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(content), data_only=True)
    formula_wb = load_workbook(BytesIO(content), data_only=False)
    changed = False
    for sheet_index, ws in enumerate(wb.worksheets):
        formula_ws = formula_wb.worksheets[sheet_index]
        # Pandas/openpyxl reads cached formula results. Workbooks created by
        # openpyxl and several export tools legitimately contain formulas
        # without a cached result; data_only=True exposes those cells as None.
        # Preserve the expression as searchable source information instead of
        # silently dropping the cell. Iterate stored cells only so a phantom
        # style range extending to XFB does not become a 16k-column scan.
        for (row, col), formula_cell in formula_ws._cells.items():
            formula = formula_cell.value
            if (
                ws.cell(row=row, column=col).value is None
                and isinstance(formula, str)
                and formula.startswith("=")
            ):
                target = ws.cell(row=row, column=col)
                target.value = formula
                # openpyxl normally interprets a leading '=' assignment as a
                # formula and pandas reopens it with data_only=True, losing it
                # again. Persist the fallback as literal source text.
                target.data_type = "s"
                changed = True
        if not ws.merged_cells.ranges:
            continue
        for merge_range in list(ws.merged_cells.ranges):
            master_value = ws.cell(merge_range.min_row, merge_range.min_col).value
            ws.unmerge_cells(str(merge_range))
            for row in range(merge_range.min_row, merge_range.max_row + 1):
                for col in range(merge_range.min_col, merge_range.max_col + 1):
                    ws.cell(row, col).value = master_value
            changed = True

    formula_wb.close()
    if not changed:
        wb.close()
        return content

    out = BytesIO()
    wb.save(out)
    wb.close()
    logger.info("Prepared merged cells and uncached formulas in XLSX before parse")
    return out.getvalue()
