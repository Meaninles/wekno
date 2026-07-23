"""Run the production splitter against rich real and synthetic fixtures."""

from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
import sys
import tempfile
import time
import zipfile
from email import policy as email_policy
from email.parser import BytesParser
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from PIL import Image
from pypdf import PdfReader

from weknora_document_splitter import service


FIXTURES = {
    "doc": "largest-real.doc",
    "docx": "largest-real.docx",
    "xlsx": "largest-real.xlsx",
    "xls": "converted-xls/largest-real.xls",
    "pptx": "complex-codex-grid-26-slides.pptx",
    "ppt": "converted-ppt/complex-codex-grid-26-slides.ppt",
    "pdf": "converted-pdf/complex-codex-grid-26-slides.pdf",
    "csv": "complex-tabular-large.csv",
    "txt": "complex-structured-large.txt",
    "text": "complex-structured-large.txt",
    "md": "complex-structured-large.md",
    "markdown": "complex-structured-large.md",
    "json": "complex-nested-large.json",
    "epub": "complex-linked-large.epub",
    "mhtml": "complex-linked-large.mhtml",
    "jpg": "complex-high-resolution.jpg",
    "jpeg": "complex-high-resolution.jpeg",
    "png": "complex-high-resolution.png",
    "webp": "complex-high-resolution.webp",
    "gif": "complex-multiframe.gif",
    "bmp": "complex-high-resolution.bmp",
    "tiff": "complex-high-resolution.tiff",
    "mp3": "complex-timeline.mp3",
    "m4a": "complex-timeline.m4a",
    "ogg": "complex-timeline.ogg",
    "flac": "complex-timeline.flac",
    "wav": "complex-timeline-large.wav",
}

EXPECTED_KINDS = {
    **{ext: {"pages"} for ext in ("doc", "docx", "ppt", "pptx", "pdf")},
    **{ext: {"sheet_range"} for ext in ("xls", "xlsx")},
    "csv": {"record_range"},
    **{ext: {"line_range"} for ext in ("txt", "text", "md", "markdown")},
    "json": {"json_path_records", "json_items"},
    "epub": {"spine_items"},
    "mhtml": {"dom_units"},
    **{
        ext: {"image_tile"}
        for ext in ("jpg", "jpeg", "png", "webp", "gif", "bmp", "tiff")
    },
    **{ext: {"time_range"} for ext in ("mp3", "m4a", "ogg", "flac", "wav")},
}


def _sha256(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _validate_physical_payload(file_type: str, data: bytes) -> None:
    if file_type == "pdf":
        assert len(PdfReader(io.BytesIO(data), strict=True).pages) > 0
        return
    if file_type == "xlsx":
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=False)
        try:
            assert len(workbook.worksheets) == 1
            assert next(workbook.active.iter_rows(values_only=True), None) is not None
        finally:
            workbook.close()
        return
    if file_type == "csv":
        decoded = data.decode("utf-8")
        assert next(csv.reader(io.StringIO(decoded)), None) is not None
        return
    if file_type in {"txt", "md"}:
        assert data.decode("utf-8").strip()
        return
    if file_type == "json":
        assert json.loads(data.decode("utf-8"))
        return
    if file_type == "epub":
        with zipfile.ZipFile(io.BytesIO(data)) as archive:
            assert archive.testzip() is None
            assert "mimetype" in archive.namelist()
        return
    if file_type == "mhtml":
        message = BytesParser(policy=email_policy.default).parsebytes(data)
        assert any(
            part.get_content_type() in {"text/html", "text/plain"}
            for part in message.walk()
        )
        return
    if file_type == "png":
        with Image.open(io.BytesIO(data)) as image:
            image.verify()
        return
    if file_type == "mp3":
        assert data.startswith(b"ID3") or data[:2] in {
            b"\xff\xfb",
            b"\xff\xf3",
            b"\xff\xf2",
        }
        return
    if file_type == "m4a":
        assert b"ftyp" in data[:64]
        return
    if file_type == "ogg":
        assert data.startswith(b"OggS")
        return
    if file_type == "flac":
        assert data.startswith(b"fLaC")
        return
    if file_type == "wav":
        assert data.startswith(b"RIFF") and data[8:12] == b"WAVE"
        return
    raise AssertionError(f"unvalidated physical file type: {file_type}")


def _validate_ordered_coverage(ext: str, parts: list[dict[str, Any]]) -> None:
    locators = [part["locator"] for part in parts]
    if ext in {"doc", "docx", "ppt", "pptx", "pdf"}:
        assert locators[0]["page_start"] == 1
        for previous, current in zip(locators, locators[1:], strict=False):
            assert current["page_start"] == previous["page_end"] + 1
        return
    if ext in {"txt", "text", "md", "markdown"}:
        assert locators[0]["line_start"] == 1
        for previous, current in zip(locators, locators[1:], strict=False):
            assert current["line_start"] <= previous["line_end"] + 1
        return
    if ext == "csv":
        groups: dict[tuple[int, int], list[dict[str, Any]]] = {}
        for locator in locators:
            key = (locator["column_start"], locator["column_end"])
            groups.setdefault(key, []).append(locator)
        for windows in groups.values():
            assert windows[0]["row_start"] == 1
            for previous, current in zip(windows, windows[1:], strict=False):
                assert current["row_start"] == previous["row_end"] + 1
        return
    if ext in {"xls", "xlsx"}:
        groups: dict[tuple[str, int, int], list[dict[str, Any]]] = {}
        for locator in locators:
            key = (
                locator["sheet"],
                locator["column_start"],
                locator["column_end"],
            )
            groups.setdefault(key, []).append(locator)
        for windows in groups.values():
            # Real workbooks may contain leading/trailing or inter-table blank
            # rows. They carry no parser content, so locators preserve their
            # source coordinates while physical windows begin at the first
            # semantic row.
            assert windows[0]["row_start"] >= 1
            if windows[0].get("header_row") == windows[0]["row_start"]:
                assert not windows[0]["header_repeated"]
            for previous, current in zip(windows, windows[1:], strict=False):
                assert current["row_start"] >= previous["row_end"] + 1
        return
    if ext == "json":
        key = (
            "record_start"
            if locators[0]["kind"] == "json_path_records"
            else "item_start"
        )
        end_key = "record_end" if key == "record_start" else "item_end"
        assert locators[0][key] == 0
        for previous, current in zip(locators, locators[1:], strict=False):
            assert current[key] == previous[end_key] + 1
        return
    if ext == "epub":
        assert locators[0]["spine_start"] == 0
        for previous, current in zip(locators, locators[1:], strict=False):
            assert current["spine_start"] == previous["spine_end"] + 1
        return
    if ext == "mhtml":
        text_locators = [
            locator
            for locator in locators
            if not locator.get("resource_continuation")
        ]
        assert text_locators[0]["unit_start"] == 0
        for previous, current in zip(
            text_locators, text_locators[1:], strict=False
        ):
            if current["unit_start"] == previous["unit_end"]:
                assert current.get("segment_start", 0) > previous.get(
                    "segment_end", -1
                )
            else:
                assert current["unit_start"] == previous["unit_end"] + 1
        covered_start = text_locators[0]["unit_start"]
        covered_end = text_locators[-1]["unit_end"]
        for locator in locators:
            if locator.get("resource_continuation"):
                assert covered_start <= locator["unit_start"] <= covered_end
        resource_groups: dict[int, list[dict[str, Any]]] = {}
        for locator in locators:
            if locator.get("resource_continuation"):
                resource_groups.setdefault(locator["resource_index"], []).append(
                    locator
                )
        for resource_locators in resource_groups.values():
            if resource_locators[0].get("resource_preserved"):
                assert len(resource_locators) == 1
            else:
                expected = resource_locators[0]["resource_tile_count"]
                assert len(resource_locators) == expected
                assert {
                    locator["resource_tile_index"]
                    for locator in resource_locators
                } == set(range(expected))
        return
    if ext in {"mp3", "m4a", "ogg", "flac", "wav"}:
        assert abs(float(locators[0]["start_seconds"])) < 0.001
        for previous, current in zip(locators, locators[1:], strict=False):
            # Adjacent audio windows intentionally overlap, but never leave a gap.
            assert current["start_seconds"] <= previous["end_seconds"]
        return
    if ext in {"jpg", "jpeg", "png", "webp", "gif", "bmp", "tiff"}:
        for locator in locators:
            assert 0 <= locator["x_start"] < locator["x_end"] <= locator["source_width"]
            assert 0 <= locator["y_start"] < locator["y_end"] <= locator["source_height"]
            assert 1 <= locator["frame_index"] <= locator["frame_count"]


def smoke_one(root: Path, ext: str, relative_path: str) -> dict[str, Any]:
    source = root / relative_path
    assert source.is_file(), source
    with tempfile.TemporaryDirectory(prefix=f"split-{ext}-") as temp:
        temp_root = Path(temp)
        archive_path = temp_root / "result.zip"
        manifest = service.create_split_archive(
            source_path=source,
            archive_path=archive_path,
            file_name=f"quality-fixture.{ext}",
            file_type=ext,
            source_size=source.stat().st_size,
            source_sha256=service._sha256_file(source),
            minimum_parts=3,
            target_ratio=0.75,
            policy=service.SplitPolicy(max_parts=10_000),
        )
        assert manifest["source"]["file_type"] == ext
        assert manifest["part_count"] >= 2
        assert len(manifest["parts"]) == manifest["part_count"]
        assert all(
            part["locator"]["kind"] in EXPECTED_KINDS[ext]
            for part in manifest["parts"]
        )
        _validate_ordered_coverage(ext, manifest["parts"])

        max_part_bytes = 0
        with zipfile.ZipFile(archive_path) as archive:
            assert archive.testzip() is None
            assert archive.getinfo("manifest.json").compress_type == zipfile.ZIP_STORED
            archive_manifest = json.loads(archive.read("manifest.json"))
            assert archive_manifest == manifest
            for part in manifest["parts"]:
                member = f"parts/{part['file_name']}"
                info = archive.getinfo(member)
                assert info.compress_type == zipfile.ZIP_STORED
                data = archive.read(member)
                assert len(data) == part["size_bytes"] == info.file_size
                assert _sha256(data) == part["sha256"]
                assert len(data) <= service._HARD_BYTES[part["file_type"]]
                assert isinstance(part["locator"], dict) and part["locator"]
                assert isinstance(part["metrics"], dict) and part["metrics"]
                _validate_physical_payload(part["file_type"], data)
                max_part_bytes = max(max_part_bytes, len(data))

        return {
            "format": ext,
            "source": relative_path,
            "source_bytes": source.stat().st_size,
            "strategy": manifest["strategy"],
            "parts": manifest["part_count"],
            "max_part_bytes": max_part_bytes,
            "physical_types": sorted(
                {part["file_type"] for part in manifest["parts"]}
            ),
            "first_locator": manifest["parts"][0]["locator"],
            "last_locator": manifest["parts"][-1]["locator"],
        }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("fixture_root", type=Path)
    parser.add_argument("--only", nargs="*", choices=sorted(FIXTURES))
    args = parser.parse_args()
    root = args.fixture_root.resolve()
    formats = args.only or list(FIXTURES)
    results = []
    for position, ext in enumerate(formats, start=1):
        started = time.monotonic()
        print(
            f"[{position}/{len(formats)}] split-smoke start: {ext}",
            file=sys.stderr,
            flush=True,
        )
        result = smoke_one(root, ext, FIXTURES[ext])
        results.append(result)
        print(
            (
                f"[{position}/{len(formats)}] split-smoke pass: {ext}; "
                f"parts={result['parts']}; elapsed={time.monotonic() - started:.2f}s"
            ),
            file=sys.stderr,
            flush=True,
        )
    print(json.dumps(results, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
