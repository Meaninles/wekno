import csv
import hashlib
import io
import json
import random
import tempfile
import unittest
import zipfile
from email import policy as email_policy
from email.message import EmailMessage
from email.parser import BytesParser
from pathlib import Path
from unittest import mock

from PIL import Image
from ebooklib import epub
from openpyxl import Workbook, load_workbook
from pypdf import PdfReader, PdfWriter

from docreader.proto import docreader_pb2
from weknora_document_splitter import grpc_adapter
from weknora_document_splitter import service


class DocumentSplitterQualityTest(unittest.TestCase):
    def test_grpc_stream_preserves_xlsx_suffix_for_format_sensitive_readers(self):
        with tempfile.TemporaryDirectory() as temp:
            source = Path(temp) / "streamed.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["编号", "名称"])
            for index in range(80):
                sheet.append([index, f"资产-{index}"])
            workbook.save(source)
            data = source.read_bytes()
            frames = [
                docreader_pb2.SplitRequest(
                    header=docreader_pb2.SplitHeader(
                        file_name=source.name,
                        file_type="xlsx",
                        source_size=len(data),
                        source_sha256=hashlib.sha256(data).hexdigest(),
                        minimum_parts=2,
                        target_ratio=0.75,
                    )
                ),
                docreader_pb2.SplitRequest(data=data),
            ]

            class ActiveContext:
                @staticmethod
                def is_active() -> bool:
                    return True

            responses = list(grpc_adapter.split_rpc(iter(frames), ActiveContext()))
            self.assertGreaterEqual(len(responses), 2)
            self.assertEqual(responses[0].WhichOneof("payload"), "header")
            self.assertGreaterEqual(responses[0].header.part_count, 2)
            self.assertNotIn(
                "error",
                [response.WhichOneof("payload") for response in responses],
            )

    def test_xlsx_leading_blank_rows_do_not_duplicate_semantic_header(self) -> None:
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            source = root / "leading-blank.xlsx"
            output = root / "parts"
            output.mkdir()
            workbook = Workbook()
            sheet = workbook.active
            sheet.append([None, None])
            sheet.append(["编号", "名称"])
            for index in range(30):
                sheet.append([index, f"资产-{index}"])
            workbook.save(source)

            parts = service._split_xlsx(
                source,
                output,
                3,
                0.75,
                service.SplitPolicy(max_parts=100),
            )
            first = parts[0]
            self.assertEqual(first.locator["header_row"], 2)
            self.assertFalse(first.locator["header_repeated"])
            parsed = load_workbook(first.path, read_only=True, data_only=True)
            try:
                rows = list(parsed.active.iter_rows(values_only=True))
            finally:
                parsed.close()
            self.assertEqual(rows.count(("编号", "名称")), 1)

    def test_text_long_line_is_split_without_data_loss(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "large.txt"
            source_text = "第一句。" + ("超长内容，" * 900) + "最后一句。\n"
            source.write_text(source_text, encoding="utf-8")
            output = root / "parts"
            output.mkdir()
            with mock.patch.dict(service._HARD_BYTES, {"txt": 2048}):
                parts = service._split_text(source, output, "txt", 4, 0.75)
            self.assertGreater(len(parts), 1)
            self.assertEqual(
                "".join(part.path.read_text(encoding="utf-8") for part in parts),
                source_text,
            )
            self.assertTrue(
                all(part.path.stat().st_size <= 2048 for part in parts)
            )
            self.assertEqual(parts[0].locator["line_start"], 1)
            self.assertEqual(parts[-1].locator["line_end"], 1)

    def test_markdown_continuations_retain_heading_context(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "large.md"
            source.write_text(
                "# 总则\n\n## 适用范围\n\n"
                + "\n".join(f"第 {index} 条：" + "规则内容。" * 20 for index in range(80)),
                encoding="utf-8",
            )
            output = root / "parts"
            output.mkdir()
            with mock.patch.dict(service._HARD_BYTES, {"md": 2048}):
                parts = service._split_text(source, output, "md", 4, 0.75)
            self.assertGreater(len(parts), 1)
            continuation = parts[1].path.read_text(encoding="utf-8")
            self.assertTrue(continuation.startswith("# 总则\n## 适用范围"))
            self.assertGreater(parts[1].locator["line_start"], 1)

    def test_csv_parts_are_valid_and_carry_schema_coordinates(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "large.csv"
            with source.open("w", encoding="utf-8", newline="") as stream:
                writer = csv.writer(stream)
                writer.writerow(["编号", "部门", "资产名称", "状态"])
                for index in range(400):
                    writer.writerow([index, f"部门{index % 9}", f"资产{index}", "有效"])
            output = root / "parts"
            output.mkdir()
            with mock.patch.dict(service._HARD_BYTES, {"csv": 4096}):
                parts = service._split_csv(source, output, 4, 0.75)
            self.assertGreater(len(parts), 1)
            self.assertFalse(parts[0].locator["header_repeated"])
            self.assertTrue(any(p.locator["header_repeated"] for p in parts[1:]))
            self.assertTrue(
                all("header_context" in part.locator for part in parts)
            )
            for part in parts:
                with part.path.open("r", encoding="utf-8", newline="") as stream:
                    self.assertGreater(len(list(csv.reader(stream))), 0)

    def test_xlsx_parts_preserve_sheets_headers_and_row_ranges(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "large.xlsx"
            workbook = Workbook()
            first = workbook.active
            first.title = "资产总表"
            first.append(["编号", "部门", "资产名称"])
            for index in range(360):
                first.append([index, f"部门{index % 7}", f"资产{index}"])
            second = workbook.create_sheet("字段字典")
            second.append(["字段", "含义"])
            for index in range(60):
                second.append([f"field_{index}", f"字段含义 {index}"])
            workbook.save(source)
            output = root / "parts"
            output.mkdir()
            with mock.patch.dict(service._HARD_BYTES, {"xlsx": 64 * 1024}):
                policy = service.SplitPolicy(max_parts=100)
                parts = service._split_xlsx(source, output, 5, 0.75, policy)
            self.assertGreaterEqual(len(parts), 5)
            self.assertEqual(
                {part.locator["sheet"] for part in parts},
                {"资产总表", "字段字典"},
            )
            self.assertTrue(
                any(part.locator["header_repeated"] for part in parts)
            )
            for part in parts:
                parsed = load_workbook(part.path, read_only=True)
                try:
                    self.assertEqual(len(parsed.worksheets), 1)
                    # Write-only workbooks intentionally omit cached sheet
                    # dimensions, so read-only openpyxl may report max_row=None.
                    # Iterating the stream is the parser-relevant assertion.
                    self.assertIsNotNone(
                        next(parsed.active.iter_rows(values_only=True), None)
                    )
                finally:
                    parsed.close()

    def test_wide_tables_repeat_keys_use_row_major_order_and_materialize_merges(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)

            csv_source = root / "wide.csv"
            csv_header = ["主键", "部门", "对象"] + [
                f"指标_{index:03d}" for index in range(4, 111)
            ]
            with csv_source.open("w", encoding="utf-8", newline="") as stream:
                writer = csv.writer(stream)
                writer.writerow(csv_header)
                for row_index in range(1, 25):
                    writer.writerow(
                        [f"ID-{row_index}", f"部门-{row_index % 3}", f"对象-{row_index}"]
                        + [f"V-{row_index}-{column}" for column in range(4, 111)]
                    )
            csv_output = root / "wide-csv-parts"
            csv_output.mkdir()
            csv_parts = service._split_csv(
                csv_source, csv_output, 4, 0.75
            )
            csv_order = [
                (
                    part.locator["row_start"],
                    part.locator["column_start"],
                )
                for part in csv_parts
            ]
            self.assertEqual(csv_order, sorted(csv_order))
            later_csv_window = next(
                part
                for part in csv_parts
                if part.locator["column_start"] > 1
            )
            self.assertEqual(
                later_csv_window.locator["anchor_columns"], [1, 2, 3]
            )
            with later_csv_window.path.open(
                "r", encoding="utf-8", newline=""
            ) as stream:
                first_row = next(csv.reader(stream))
            self.assertEqual(first_row[:3], csv_header[:3])

            xlsx_source = root / "wide.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "合并宽表"
            xlsx_header = ["主键", "部门", "对象"] + [
                f"指标_{index:03d}" for index in range(4, 111)
            ]
            sheet.append(xlsx_header)
            for row_index in range(1, 25):
                sheet.append(
                    [f"ID-{row_index}", f"部门-{row_index % 3}", f"对象-{row_index}"]
                    + [f"V-{row_index}-{column}" for column in range(4, 111)]
                )
            sheet["B2"] = "合并部门"
            sheet.merge_cells("B2:B4")
            workbook.save(xlsx_source)
            xlsx_output = root / "wide-xlsx-parts"
            xlsx_output.mkdir()
            with mock.patch.dict(service._HARD_BYTES, {"xlsx": 64 * 1024}):
                xlsx_parts = service._split_xlsx(
                    xlsx_source,
                    xlsx_output,
                    4,
                    0.75,
                    service.SplitPolicy(max_parts=100),
                )
            xlsx_order = [
                (
                    part.locator["row_start"],
                    part.locator["column_start"],
                )
                for part in xlsx_parts
            ]
            self.assertEqual(xlsx_order, sorted(xlsx_order))
            later_xlsx_window = next(
                part
                for part in xlsx_parts
                if part.locator["column_start"] > 1
            )
            self.assertEqual(
                later_xlsx_window.locator["anchor_columns"], [1, 2, 3]
            )
            parsed = load_workbook(
                later_xlsx_window.path, read_only=True, data_only=True
            )
            try:
                first_row = next(parsed.active.iter_rows(values_only=True))
            finally:
                parsed.close()
            self.assertEqual(tuple(first_row[:3]), tuple(xlsx_header[:3]))

            merged_part = next(
                part
                for part in xlsx_parts
                if part.locator["column_start"] == 1
                and part.locator["row_start"] <= 2
                and part.locator["row_end"] >= 4
            )
            self.assertTrue(
                merged_part.locator["merged_values_materialized"]
            )
            parsed = load_workbook(
                merged_part.path, read_only=True, data_only=True
            )
            try:
                values = list(parsed.active.iter_rows(values_only=True))
            finally:
                parsed.close()
            self.assertEqual(
                [values[index][1] for index in range(1, 4)],
                ["合并部门", "合并部门", "合并部门"],
            )

    def test_json_array_ndjson_and_nested_object_parts_are_valid(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            with mock.patch.dict(service._HARD_BYTES, {"json": 2048}):
                top_array = root / "array.json"
                top_array.write_text(
                    json.dumps(
                        [{"id": index, "name": f"资产{index}"} for index in range(250)],
                        ensure_ascii=False,
                    ),
                    encoding="utf-8",
                )
                array_output = root / "array-parts"
                array_output.mkdir()
                array_parts = service._split_json(
                    top_array, array_output, 4, 0.75
                )
                self.assertGreater(len(array_parts), 1)
                self.assertEqual(
                    sum(len(json.loads(p.path.read_text("utf-8"))) for p in array_parts),
                    250,
                )

                ndjson = root / "records.json"
                ndjson.write_text(
                    "\n".join(
                        json.dumps({"id": index, "value": f"记录{index}"}, ensure_ascii=False)
                        for index in range(200)
                    ),
                    encoding="utf-8",
                )
                ndjson_output = root / "ndjson-parts"
                ndjson_output.mkdir()
                ndjson_parts = service._split_json(
                    ndjson, ndjson_output, 4, 0.75
                )
                self.assertTrue(
                    all(p.locator.get("content_kind") == "ndjson" for p in ndjson_parts)
                )
                self.assertEqual(
                    sum(len(json.loads(p.path.read_text("utf-8"))) for p in ndjson_parts),
                    200,
                )

                nested = root / "nested.json"
                nested.write_text(
                    json.dumps(
                        {
                            "metadata": {"system": "资产平台"},
                            "data": [
                                {"id": index, "name": f"数据资产{index}"}
                                for index in range(350)
                            ],
                        },
                        ensure_ascii=False,
                    ),
                    encoding="utf-8",
                )
                nested_output = root / "nested-parts"
                nested_output.mkdir()
                nested_parts = service._split_json(
                    nested, nested_output, 4, 0.75
                )
                self.assertTrue(
                    all(p.locator["kind"] == "json_path_records" for p in nested_parts)
                )
                records = []
                for part in nested_parts:
                    payload = json.loads(part.path.read_text("utf-8"))
                    records.extend(payload["__weknora_path_records"])
                paths = {record["path"] for record in records}
                self.assertIn('$["metadata"]["system"]', paths)
                self.assertIn('$["data"][349]["name"]', paths)

    def test_pdf_pages_and_image_frames_preserve_coordinates(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            pdf = root / "pages.pdf"
            writer = PdfWriter()
            for index in range(8):
                writer.add_blank_page(width=600 + index, height=800)
            with pdf.open("wb") as stream:
                writer.write(stream)
            pdf_output = root / "pdf-parts"
            pdf_output.mkdir()
            pdf_parts = service._split_pdf(pdf, pdf_output, 4, 0.75)
            self.assertEqual(len(pdf_parts), 4)
            self.assertEqual(pdf_parts[0].locator["page_start"], 1)
            self.assertEqual(pdf_parts[-1].locator["page_end"], 8)
            self.assertEqual(
                sum(len(PdfReader(str(part.path)).pages) for part in pdf_parts),
                8,
            )

            image_path = root / "frames.gif"
            frames = [
                Image.new("RGB", (320, 240), color)
                for color in ("red", "green", "blue")
            ]
            frames[0].save(
                image_path,
                save_all=True,
                append_images=frames[1:],
                duration=100,
                loop=0,
            )
            image_output = root / "image-parts"
            image_output.mkdir()
            policy = service.SplitPolicy(image_tile_pixels=40_000, max_parts=100)
            image_parts = service._split_image(
                image_path, image_output, "gif", 3, 0.75, policy
            )
            self.assertEqual(
                {part.locator["frame_index"] for part in image_parts},
                {1, 2, 3},
            )
            self.assertTrue(
                all(part.locator["frame_count"] == 3 for part in image_parts)
            )
            self.assertTrue(
                all(
                    part.locator["x_start"] == 0
                    and part.locator["x_end"] == 320
                    for part in image_parts
                )
            )

    def test_raw_tiff_is_losslessly_normalized_before_byte_based_tiling(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "raw.tiff"
            image = Image.new("RGB", (1600, 900), "white")
            image.save(source, format="TIFF", compression="raw")
            self.assertGreater(source.stat().st_size, service._HARD_BYTES["tiff"])

            output = root / "parts"
            output.mkdir()
            parts = service._split_image(
                source,
                output,
                "tiff",
                minimum_parts=3,
                ratio=0.75,
                policy=service.SplitPolicy(),
            )

            self.assertEqual(len(parts), 1)
            part = parts[0]
            self.assertEqual(part.file_type, "png")
            self.assertLessEqual(
                part.path.stat().st_size, service._HARD_BYTES["png"]
            )
            self.assertEqual(
                part.locator,
                {
                    "kind": "image_tile",
                    "frame_index": 1,
                    "frame_count": 1,
                    "x_start": 0,
                    "y_start": 0,
                    "x_end": 1600,
                    "y_end": 900,
                    "source_width": 1600,
                    "source_height": 900,
                },
            )
            self.assertTrue(part.metrics["normalized_whole_frame"])

    def test_epub_and_mhtml_do_not_duplicate_unreferenced_images(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            image_a = self._png_bytes("red")
            image_b = self._png_bytes("blue")

            book = epub.EpubBook()
            book.set_identifier("quality-test")
            book.set_title("Quality")
            book.set_language("zh")
            chapter_a = epub.EpubHtml(
                uid="a", file_name="chapters/a.xhtml", title="A"
            )
            chapter_a.set_content(
                b"<html><body><h1>A</h1><img src='../images/a.png'/></body></html>"
            )
            chapter_b = epub.EpubHtml(
                uid="b", file_name="chapters/b.xhtml", title="B"
            )
            chapter_b.set_content(
                b"<html><body><h1>B</h1><img src='../images/b.png'/></body></html>"
            )
            resource_a = epub.EpubImage(
                uid="img-a", file_name="images/a.png", media_type="image/png"
            )
            resource_a.set_content(image_a)
            resource_b = epub.EpubImage(
                uid="img-b", file_name="images/b.png", media_type="image/png"
            )
            resource_b.set_content(image_b)
            for item in (chapter_a, chapter_b, resource_a, resource_b):
                book.add_item(item)
            book.spine = [chapter_a, chapter_b]
            book.toc = (chapter_a, chapter_b)
            book.add_item(epub.EpubNcx())
            book.add_item(epub.EpubNav())
            epub_path = root / "book.epub"
            epub.write_epub(str(epub_path), book)
            epub_output = root / "epub-parts"
            epub_output.mkdir()
            epub_parts = service._split_epub(epub_path, epub_output, 2, 0.75)
            self.assertEqual(len(epub_parts), 2)
            with zipfile.ZipFile(epub_parts[0].path) as archive:
                names = archive.namelist()
                self.assertTrue(any(name.endswith("images/a.png") for name in names))
                self.assertFalse(any(name.endswith("images/b.png") for name in names))
            with zipfile.ZipFile(epub_parts[1].path) as archive:
                names = archive.namelist()
                self.assertTrue(any(name.endswith("images/b.png") for name in names))
                self.assertFalse(any(name.endswith("images/a.png") for name in names))

            message = EmailMessage()
            message["Subject"] = "Archive"
            message.make_related()
            body = EmailMessage()
            body.set_content(
                "<html><body>"
                "<section><h1>A</h1><p>"
                + ("正文A" * 9_000)
                + "</p><img src='cid:image-a'/></section>"
                "<section><h1>B</h1><p>"
                + ("正文B" * 9_000)
                + "</p><img src='cid:image-b'/></section>"
                "</body></html>",
                subtype="html",
            )
            message.attach(body)
            message.attach(self._mhtml_image(image_a, "image-a", "a.png"))
            message.attach(self._mhtml_image(image_b, "image-b", "b.png"))
            mhtml = root / "archive.mhtml"
            mhtml.write_bytes(message.as_bytes(policy=email_policy.default))
            mhtml_output = root / "mhtml-parts"
            mhtml_output.mkdir()
            with mock.patch.dict(service._HARD_BYTES, {"mhtml": 128 * 1024}):
                mhtml_parts = service._split_mhtml(
                    mhtml, mhtml_output, 2, 0.75
                )
            payloads = [part.path.read_bytes() for part in mhtml_parts]
            self.assertTrue(all(len(payload) <= 128 * 1024 for payload in payloads))
            self.assertEqual(sum(payload.count(b"<image-a>") for payload in payloads), 1)
            self.assertEqual(sum(payload.count(b"<image-b>") for payload in payloads), 1)
            html_text = "".join(
                mime_part.get_content()
                for payload in payloads
                for mime_part in BytesParser(
                    policy=email_policy.default
                ).parsebytes(payload).walk()
                if mime_part.get_content_type() == "text/html"
            )
            self.assertEqual(html_text.count("正文A"), 9_000)
            self.assertEqual(html_text.count("正文B"), 9_000)
            resource_parts = [
                part
                for part in mhtml_parts
                if part.locator.get("resource_continuation")
            ]
            self.assertEqual(len(resource_parts), 2)
            self.assertTrue(
                all(
                    part.locator.get("resource_preserved")
                    for part in resource_parts
                )
            )

    def test_mhtml_oversized_image_is_losslessly_tiled_once(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            random_bytes = random.Random(20260724).randbytes(512 * 512 * 3)
            image_buffer = io.BytesIO()
            Image.frombytes("RGB", (512, 512), random_bytes).save(
                image_buffer, format="PNG"
            )

            message = EmailMessage()
            message["Subject"] = "Oversized embedded resource"
            message.make_related()
            body = EmailMessage()
            body.set_content(
                "<html><body><section><h1>Oversized</h1>"
                "<p>MHTML-UNIQUE-BODY-5A91</p>"
                "<img src='cid:large-noise' alt='complex noise'/></section>"
                "</body></html>",
                subtype="html",
            )
            message.attach(body)
            message.attach(
                self._mhtml_image(
                    image_buffer.getvalue(), "large-noise", "large-noise.png"
                )
            )
            source = root / "oversized.mhtml"
            source.write_bytes(message.as_bytes(policy=email_policy.default))
            output = root / "parts"
            output.mkdir()

            with mock.patch.dict(service._HARD_BYTES, {"mhtml": 96 * 1024}):
                parts = service._split_mhtml(source, output, 2, 0.75)

            self.assertTrue(
                all(part.path.stat().st_size <= 96 * 1024 for part in parts)
            )
            tiles = [
                part
                for part in parts
                if part.locator.get("resource_continuation")
            ]
            self.assertGreater(len(tiles), 1)
            self.assertTrue(
                all(not part.locator["resource_preserved"] for part in tiles)
            )
            self.assertEqual(
                {part.locator["resource_tile_index"] for part in tiles},
                set(range(len(tiles))),
            )
            self.assertTrue(
                all(
                    part.locator["resource_tile_count"] == len(tiles)
                    for part in tiles
                )
            )
            html_text = "".join(
                mime_part.get_content()
                for part in parts
                for mime_part in BytesParser(
                    policy=email_policy.default
                ).parsebytes(part.path.read_bytes()).walk()
                if mime_part.get_content_type() == "text/html"
            )
            self.assertEqual(html_text.count("MHTML-UNIQUE-BODY-5A91"), 1)

    @staticmethod
    def _png_bytes(color: str) -> bytes:
        output = io.BytesIO()
        Image.new("RGB", (64, 64), color).save(output, format="PNG")
        return output.getvalue()

    @staticmethod
    def _mhtml_image(data: bytes, content_id: str, filename: str) -> EmailMessage:
        part = EmailMessage()
        part.set_content(data, maintype="image", subtype="png")
        part["Content-ID"] = f"<{content_id}>"
        part["Content-Location"] = filename
        return part


if __name__ == "__main__":
    unittest.main()
